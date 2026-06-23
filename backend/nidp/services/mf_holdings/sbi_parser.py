"""SBI Mutual Fund monthly portfolio Excel parser.

SEBI Circular SEBI/HO/IMD/DF3/CIR/P/2020/197 mandates portfolio disclosures
in a structured format. SBI MF's Excel typically has one sheet per scheme
(or named sections within a single sheet) with columns:

  Name of Instrument | ISIN | Industry/Sector | Rating | Quantity |
  Market/Fair Value (Rs in Lakhs) | % to Net Assets

Debt-fund disclosures additionally include:
  Maturity Date | Coupon Rate | YTM (%)

The scheme name appears as a merged-cell header above each block.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Any, Optional

import openpyxl

try:
    import xlrd as _xlrd  # type: ignore[import-untyped]
    _XLRD_AVAILABLE = True
except ImportError:
    _XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)

SOURCE_TAG = "SBI_MF_PORTFOLIO_XLSX"   # default; callers may override

_COL_ALIASES: list[tuple[str, str]] = [
    ("name of instrument", "security_name"),   # SBI / most AMCs
    ("name of the instrument", "security_name"),  # Nippon
    ("instrument name",    "security_name"),
    ("security name",      "security_name"),
    ("isin",               "security_isin"),
    ("industry",           "sector"),
    ("sector",             "sector"),
    ("rating",             "rating"),
    ("quantity",           "quantity"),
    ("market value",       "market_value_inr"),
    ("fair value",         "market_value_inr"),
    ("% to net assets",    "weight_pct"),
    ("% net assets",       "weight_pct"),
    ("% of net",           "weight_pct"),
    ("% to nav",           "weight_pct"),      # Nippon
    ("% to aum",           "weight_pct"),
    # Debt-fund specific columns (SEBI circular format)
    ("maturity date",      "maturity_date"),   # "Maturity Date", "Maturity  Date"
    ("residual maturity",  "maturity_date"),   # some AMCs use "Residual Maturity"
    ("redemption date",    "maturity_date"),   # alternate naming
    ("ytm",                "ytm_pct"),         # "YTM", "YTM (%)", "YTM(%)"
    ("yield to maturity",  "ytm_pct"),
    ("yield",              "ytm_pct"),         # "Yield (%)" — catch-all
]

# Date format patterns used in AMFI disclosures (try in order)
_DATE_FORMATS = [
    "%d-%m-%Y",    # 31-03-2027
    "%d/%m/%Y",    # 31/03/2027
    "%d-%b-%Y",    # 31-Mar-2027
    "%d/%b/%Y",    # 31/Mar/2027
    "%d-%B-%Y",    # 31-March-2027
    "%d %b %Y",    # 31 Mar 2027  (after ordinal-suffix strip: "31st Mar 2027" → "31 Mar 2027")
    "%d %B %Y",    # 31 March 2027
    "%Y-%m-%d",    # 2027-03-31 (ISO)
]


def _to_date(v: Any) -> Optional[date]:
    """Parse a cell value to a date. Returns None on failure."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s in {"-", "N.A.", "NA", "—", "N/A"}:
        return None
    # Strip ordinal suffixes: "31st" → "31"
    s = re.sub(r"(\d+)(st|nd|rd|th)\b", r"\1", s, flags=re.IGNORECASE)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

_ITYPE_KEYWORDS: list[tuple[str, str]] = [
    ("equity shares",              "EQUITY"),
    ("equity & equity",            "EQUITY"),
    ("listed equity",              "EQUITY"),
    ("government securities",      "DEBT"),
    ("government bond",            "DEBT"),
    ("treasury bill",              "DEBT"),
    ("t-bill",                     "DEBT"),
    ("commercial paper",           "DEBT"),
    ("certificate of deposit",     "DEBT"),
    ("non-convertible debenture",  "DEBT"),
    ("ncd",                        "DEBT"),
    ("bond",                       "DEBT"),
    ("debenture",                  "DEBT"),
    ("futures",                    "DERIVATIVE"),
    ("options",                    "DERIVATIVE"),
    ("reit",                       "REIT"),
    ("invit",                      "REIT"),
    ("cash & cash equivalent",     "CASH"),
    ("cblo",                       "CASH"),
    ("reverse repo",               "CASH"),
    ("net receivable",             "CASH"),
    ("mutual fund unit",           "OTHER"),
]


def _guess_itype(name: str, sector: Optional[str]) -> Optional[str]:
    haystack = ((name or "") + " " + (sector or "")).lower()
    for kw, itype in _ITYPE_KEYWORDS:
        if kw in haystack:
            return itype
    return None


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None


def _col_idx(headers: list[str], *fragments: str) -> Optional[int]:
    for frag in fragments:
        for i, h in enumerate(headers):
            if frag in h:
                return i
    return None


def _normalise_weights(rows: list[dict]) -> list[dict]:
    """Normalise weight_pct to decimal fraction (0.0–1.0 range).

    Some AMC files (e.g. HDFC) publish weight_pct as a percentage (6.47),
    others (e.g. Nippon) as a decimal fraction (0.0421). We detect the
    scale by the maximum value: if any single security exceeds 2.0, the
    values are in percentage form and should be divided by 100.

    The 2.0 threshold is safe because no single holding legitimately
    represents >200% of a fund's NAV in decimal form (and in percentage
    form, >2% is extremely common for top holdings).
    """
    weights = [r["weight_pct"] for r in rows if r.get("weight_pct") is not None]
    if not weights:
        return rows
    if max(weights) > 2.0:
        # Percentage form → convert to decimal
        for r in rows:
            if r.get("weight_pct") is not None:
                r["weight_pct"] = r["weight_pct"] / 100.0
    return rows


def parse_sbi_multisheet_xlsx(
    data: bytes,
    as_of_month: date,
    source_url: str,
    source_tag: str = SOURCE_TAG,
) -> list[dict]:
    """Parse SBI MF multi-sheet portfolio xlsx.

    SBI publishes one sheet per base fund.  An 'Index' sheet maps:
      col 1 (Scheme Short code) → col 2 (Scheme Name).

    Each scheme sheet has:
      Row 0: empty
      Row 1: ['', '', 'SBI Mutual Fund', <code>, ...]
      Row 2: ['', '', 'SCHEME NAME :', <scheme_name>, ...]
      Row 5: header row with 'Name of the Instrument', 'ISIN', etc.
      Row 6: blank (would reset header_map in generic parser — skip here)
      Row 7+: section labels + data rows

    We detect the header once per sheet and do NOT reset on blank rows,
    since every sheet is a single scheme.
    """
    import io as _io
    as_of_str = date(as_of_month.year, as_of_month.month, 1).isoformat()
    result: list[dict] = []

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("sbi_parser: cannot open xlsx: %s", e)
        return []

    # Read Index: shortcode → full scheme name
    shortcode_to_name: dict[str, str] = {}
    if "Index" in wb.sheetnames:
        for row in list(wb["Index"].iter_rows(values_only=True))[2:]:
            if row[1] and row[2]:
                shortcode_to_name[str(row[1]).strip()] = str(row[2]).strip()

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        scheme_name = shortcode_to_name.get(sheet_name, sheet_name)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Find header row (contains 'name of instrument' / 'isin')
        header_map: dict[int, str] = {}
        header_idx = -1
        for idx, row in enumerate(all_rows):
            candidate: dict[int, str] = {}
            for i, cell in enumerate(row):
                cl = str(cell or "").lower().strip()
                for alias, field in _COL_ALIASES:
                    if alias in cl:
                        if field not in candidate.values():
                            candidate[i] = field
                        break
            if "security_name" in candidate.values():
                header_map = candidate
                header_idx = idx
                break

        if not header_map or header_idx < 0:
            continue

        isin_col = next((i for i, f in header_map.items() if f == "security_isin"), None)
        val_col  = next((i for i, f in header_map.items() if f == "market_value_inr"), None)
        name_col = next((i for i, f in header_map.items() if f == "security_name"), None)

        for row in all_rows[header_idx + 1:]:
            str_cells = [str(c or "").strip() for c in row]
            non_empty  = [c for c in str_cells if c]
            if not non_empty:
                continue  # blank row — skip but do NOT reset header

            isin_val = str_cells[isin_col] if isin_col is not None and isin_col < len(str_cells) else ""
            val_val  = str_cells[val_col]  if val_col  is not None and val_col  < len(str_cells) else ""
            if not isin_val and not val_val:
                continue  # section header / annotation

            if name_col is None or name_col >= len(str_cells) or not str_cells[name_col]:
                continue
            sec_name = str_cells[name_col]
            if sec_name.lower() in _HEADER_REPEAT_NAMES:
                continue

            def _get(field: str, _sc=str_cells, _hm=header_map) -> Optional[str]:
                for ci, fn in _hm.items():
                    if fn == field and ci < len(_sc) and _sc[ci]:
                        return _sc[ci]
                return None

            isin    = _get("security_isin")
            sector  = _get("sector")
            rating  = _get("rating")
            qty     = _to_float(_get("quantity"))
            mkt_val = _to_float(_get("market_value_inr"))
            if mkt_val is not None:
                mkt_val *= 100_000  # Lakhs → INR
            weight  = _to_float(_get("weight_pct"))
            # Debt-fund specific: maturity date and YTM
            mat_raw = next(
                (row[ci] for ci, fn in header_map.items() if fn == "maturity_date" and ci < len(row)),
                None,
            )
            mat_date = _to_date(mat_raw)
            ytm_raw  = _to_float(_get("ytm_pct"))
            if ytm_raw is not None and ytm_raw < 1.0:
                ytm_raw = round(ytm_raw * 100, 4)

            result.append({
                "scheme_code":      scheme_name,
                "as_of_month":      as_of_str,
                "security_isin":    isin if (isin and isin not in {"-", "N.A.", "NA"}) else None,
                "security_name":    sec_name,
                "instrument_type":  _guess_itype(sec_name, sector),
                "sector":           sector,
                "rating":           rating,
                "quantity":         qty,
                "market_value_inr": mkt_val,
                "weight_pct":       weight,
                "maturity_date":    mat_date,
                "ytm_pct":          ytm_raw,
                "source":           source_tag,
                "source_url":       source_url,
            })

    return _normalise_weights(result)


def parse_portfolio_xlsx(
    data: bytes,
    as_of_month: date,
    source_url: str,
    source_tag: str = SOURCE_TAG,
) -> list[dict]:
    """Parse an AMC portfolio Excel → list of mf_holdings_monthly rows.

    Handles both .xlsx (openpyxl) and legacy .xls (xlrd) formats.
    Normalises weight_pct to decimal fraction (0.0–1.0) regardless of
    AMC-specific format.
    scheme_code is set to the scheme *name* from the Excel header row.
    The calling adapter resolves name → AMFI code via DB lookup.
    """
    as_of_str = date(as_of_month.year, as_of_month.month, 1).isoformat()
    result: list[dict] = []

    # Try openpyxl first (.xlsx); fall back to xlrd for legacy .xls
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = list(ws.iter_rows(values_only=True))
            result.extend(
                _parse_sheet(sheet_rows, as_of_str, source_url, sheet_name, source_tag)
            )
    except Exception as xlsx_err:
        if not _XLRD_AVAILABLE:
            raise
        logger.debug("openpyxl failed (%s), retrying with xlrd: %s", type(xlsx_err).__name__, xlsx_err)

        # .xls fallback via xlrd
        xls_wb = _xlrd.open_workbook(file_contents=data)
        for sheet_name in xls_wb.sheet_names():
            ws = xls_wb.sheet_by_name(sheet_name)
            sheet_rows = [
                tuple(ws.cell_value(r, c) for c in range(ws.ncols))
                for r in range(ws.nrows)
            ]
            result.extend(
                _parse_sheet(sheet_rows, as_of_str, source_url, sheet_name, source_tag)
            )

    return _normalise_weights(result)


_SCHEME_NAME_KWS = ("fund", "plan", "growth", "direct", "regular", "idcw", "dividend")
# "scheme"/"folio" excluded: "scheme" matches disclaimers ("Scheme & Benchmark Riskometer..."),
# "folio" is a substring of "portfolio" — both cause false scheme-name detection.

_HEADER_REPEAT_NAMES = {
    "name of instrument", "name of the instrument",
    "instrument name", "security name",
}


_VALID_ISIN_RE = re.compile(r"^IN[EF][A-Z0-9]{9}$")
# SEBI sub-section bullets: "(a) Listed/awaiting listing", "(b) Unlisted",
# "(c) Privately placed", etc. — labels, not holdings.
_SECTION_BULLET_RE = re.compile(r"^\s*\([a-e]\)", re.I)


def _is_aggregation_row(name: str, has_valid_isin: bool) -> bool:
    """True for subtotal / grand-total / section-label rows that must not be counted
    as holdings. They carry a value+weight (so the no-data filter misses them), and
    including them double-counts — e.g. NIPPON schemes summed to ~224% because every
    section's Sub Total plus the Grand Total were ingested alongside the real holdings.
    A row with a valid ISIN is always a real security and is never treated as a total.
    """
    if has_valid_isin:
        return False
    n = name.strip()
    if "total" in n.lower():            # subtotal / sub total / grand total / total
        return True
    if _SECTION_BULLET_RE.match(n):     # "(a) Listed", "(b) Unlisted", ...
        return True
    return False


def _parse_sheet(
    rows: list[tuple],
    as_of_month: str,
    source_url: str,
    sheet_name: str,
    source_tag: str = SOURCE_TAG,
) -> list[dict]:
    result: list[dict] = []
    current_scheme: Optional[str] = None
    header_map: dict[int, str] = {}   # col_index → field_name

    # ── Pre-scan first 3 rows for scheme name ────────────────────────────
    # Handles two patterns:
    #   a) Nippon-style: [internal_code(≤12), full_scheme_name, ...]
    #   b) HDFC-style:   [fund_name, fund_name, fund_name, ...] (same in all cols)
    for row in rows[:3]:
        str_cells = [str(c or "").strip() for c in row]
        non_empty = [c for c in str_cells if c]
        if len(non_empty) >= 2:
            first, second = non_empty[0], non_empty[1]
            # Pattern (b): majority of non-empty cells share same fund name
            # (HDFC has fund name repeated across most cols; last cols may differ)
            from collections import Counter as _Counter
            most_common_val, most_common_count = _Counter(non_empty).most_common(1)[0]
            if (most_common_count >= max(2, len(non_empty) // 2)
                    and ("fund" in most_common_val.lower() or "etf" in most_common_val.lower()
                         or any(kw in most_common_val.lower() for kw in _SCHEME_NAME_KWS))):
                current_scheme = most_common_val
                break
            # Pattern (a): short alphanumeric code in first cell + fund/ETF name in second
            if (len(first) <= 12 and first.replace(" ", "").isalnum()
                    and ("fund" in second.lower() or "etf" in second.lower()
                         or any(kw in second.lower() for kw in _SCHEME_NAME_KWS))):
                current_scheme = second
                break

    for row in rows:
        str_cells = [str(c or "").strip() for c in row]
        non_empty  = [c for c in str_cells if c]
        if not non_empty:
            # blank separator — a new scheme block may follow; reset header
            if header_map:
                header_map = {}
            continue

        # ── Try to (re)build header map ──────────────────────────────
        if not header_map:
            candidate: dict[int, str] = {}
            for i, cell in enumerate(str_cells):
                cl = cell.lower()
                for alias, field in _COL_ALIASES:
                    if alias in cl:
                        if field not in candidate.values():  # first match wins
                            candidate[i] = field
                        break
            if "security_name" in candidate.values():
                header_map = candidate
                continue
            # Not a header row — could be scheme name (single-cell rows only)
            if len(non_empty) == 1:
                name = non_empty[0]
                nl = name.lower()
                if any(kw in nl for kw in _SCHEME_NAME_KWS):
                    current_scheme = name
            continue

        # ── Sub-section label detection ──────────────────────────────
        # Lines like "Equity & Equity related Instruments" with no numeric data.
        # Skip if there is no ISIN-column value and no market value.
        isin_col = next((i for i, f in header_map.items() if f == "security_isin"), None)
        val_col  = next((i for i, f in header_map.items() if f == "market_value_inr"), None)
        isin_val = str_cells[isin_col] if isin_col is not None and isin_col < len(str_cells) else ""
        val_val  = str_cells[val_col]  if val_col  is not None and val_col  < len(str_cells) else ""
        if not isin_val and not val_val:
            continue  # category header / annotation row — no actual security data

        # ── Data row ─────────────────────────────────────────────────
        name_col = next((i for i, f in header_map.items() if f == "security_name"), None)
        if name_col is None or name_col >= len(str_cells) or not str_cells[name_col]:
            continue

        sec_name = str_cells[name_col]
        if sec_name.lower() in _HEADER_REPEAT_NAMES:
            continue  # repeated header row

        def _get(field: str) -> Optional[str]:
            for ci, fn in header_map.items():
                if fn == field and ci < len(str_cells) and str_cells[ci]:
                    return str_cells[ci]
            return None

        isin    = _get("security_isin")
        # Drop subtotal / grand-total / section-label rows (they carry a value so the
        # no-data filter above lets them through, but counting them double-counts).
        if _is_aggregation_row(sec_name, bool(isin and _VALID_ISIN_RE.match(isin))):
            continue
        sector  = _get("sector")
        rating  = _get("rating")
        qty     = _to_float(_get("quantity"))
        mkt_val = _to_float(_get("market_value_inr"))
        if mkt_val is not None:
            mkt_val *= 100_000          # Lakhs → INR
        weight  = _to_float(_get("weight_pct"))
        # Debt-fund specific: maturity date and YTM
        mat_raw = next(
            (row[ci] for ci, fn in header_map.items() if fn == "maturity_date" and ci < len(row)),
            None,
        )
        mat_date = _to_date(mat_raw)
        ytm_raw  = _to_float(_get("ytm_pct"))
        # YTM may be published as a decimal (0.0721) or percentage (7.21); normalise to %
        if ytm_raw is not None and ytm_raw < 1.0:
            ytm_raw = round(ytm_raw * 100, 4)

        result.append({
            # scheme_code will be resolved by the adapter (name → AMFI code).
            "scheme_code":      current_scheme or sheet_name,
            "as_of_month":      as_of_month,
            "security_isin":    isin if (isin and isin not in {"-", "N.A.", "NA"}) else None,
            "security_name":    sec_name,
            "instrument_type":  _guess_itype(sec_name, sector),
            "sector":           sector,
            "rating":           rating,
            "quantity":         qty,
            "market_value_inr": mkt_val,
            "weight_pct":       weight,
            "maturity_date":    mat_date,
            "ytm_pct":          ytm_raw,
            "source":           source_tag,
            "source_url":       source_url,
        })
    return result
