"""AdvisorKhoj-sourced monthly portfolio holdings (full security-level).

The SEBI-mandated monthly portfolio disclosure — every security, not the
factsheet's top-10 — is the right source for holdings, but each AMC
publishes it in a different (often JS-rendered) location, which is why
the direct adapters for several big AMCs are broken. AdvisorKhoj indexes
the SAME official files in STATIC HTML at a uniform URL per AMC, so we
scrape it to find the target month's disclosure file (whose link points
back to the official AMC source), download, and parse.

This is the holdings fallback for AMCs whose direct adapter is broken
(ICICI, Kotak, ABSL, UTI, Axis …). The disclosure is a zip of per-fund
xlsx in the standard SEBI format:

  Company/Issuer/Instrument | ISIN | Industry/Rating | Quantity |
  Exposure/Market Value (Rs Lakh) | % to Nav

We keep ISIN-keyed securities only — that drops section subtotals, cash,
TREPS and derivative rows (which carry a weight but no ISIN and would
otherwise double-count), and matches the downstream ISIN-keyed primitives
(migration 058). The scheme name is the xlsx filename; one fund's
holdings fan out to all its plan-variant scheme_codes (the convention the
concentration/active-share views read).
"""
from __future__ import annotations

import calendar
import io
import logging
import re
import zipfile
from datetime import date
from typing import Optional
from urllib.parse import urljoin

import aiohttp
import openpyxl

try:
    import xlrd as _xlrd          # legacy .xls (UTI/ABSL/some others)
    _XLRD = True
except ImportError:
    _XLRD = False

from .sbi_parser import _to_float  # reuse the tolerant numeric coercion

logger = logging.getLogger(__name__)

_AK_BASE = "https://www.advisorkhoj.com/form-download-centre/Mutual"

# amc_id → AdvisorKhoj slug. Covers the broken-adapter AMCs first; extend as
# each is verified.
AK_SLUG: dict[str, str] = {
    "icici_pru": "ICICI-Prudential-Mutual-Fund",
    "kotak":     "Kotak-Mahindra-Mutual-Fund",
    "absl":      "Aditya-Birla-Sun-Life-Mutual-Fund",
    "uti":       "UTI-Mutual-Fund",
    "axis":      "Axis-Mutual-Fund",
}

_ISIN_RE = re.compile(r"^IN[EF][A-Z0-9]{9}$")
_DISCLOSURE_EXT_RE = re.compile(r"\.(?:zip|xlsx|xls)(?:$|\?)", re.I)
# Header column synonyms (substring match against the disclosure header row).
_H_NAME = ("company", "issuer", "instrument", "name of")
_H_MV = ("exposure/market", "market value", "fair value", "market/fair")
_H_WT = ("% to nav", "% to net", "% of net", "% to aum")


def _month_patterns(as_of_month: date) -> list[str]:
    """Regexes that identify as_of_month in a disclosure-file URL. AMCs encode
    it inconsistently: by month name (.../2026/May/, as-on-May-2026) or by the
    month-END date in DD?MM?YYYY / DD?MM?YY form (31.05.2026, 31_05_26, 31052026,
    .../2026/3105...). We match the month-end date because disclosures are
    'as on' the last day of the month."""
    full = as_of_month.strftime("%B").lower()      # 'may'
    abbr = as_of_month.strftime("%b").lower()      # 'may' / 'apr'
    yr, mo = as_of_month.year, as_of_month.month
    last = calendar.monthrange(yr, mo)[1]          # 31 for May
    sep = r"[._\-/ ,]?"                            # optional separator (incl. comma)
    # Month-name patterns: no leading lookbehind, so a glued 'PortfolioMay2026'
    # (Kotak) matches; the month name + nearby year is specific enough.
    # Numeric patterns use a digit lookbehind (not \b, which won't fire after '_').
    return [
        rf"/{yr}/{full}/", rf"/{yr}/{abbr}/",                                  # ICICI path
        rf"{full}{sep}(?:{last}{sep})?{yr}(?!\d)",                            # 'May2026' / 'May-31-2026'
        rf"{abbr}{sep}(?:{last}{sep})?{yr}(?!\d)",
        rf"{last:02d}{sep}{full}{sep}{yr}(?!\d)",                             # '31-May-2026' (day first)
        rf"(?<!\d){last:02d}{sep}{mo:02d}{sep}{yr}(?!\d)",                     # 31.05.2026 / 31052026
        rf"(?<!\d){last:02d}{sep}{mo:02d}{sep}{yr % 100:02d}(?!\d)",           # 31_05_26 (2-digit yr)
        rf"/{yr}{sep}{mo:02d}(?!\d)",                                         # /2026/05, 2026-05
    ]


def _file_url_for_month(html: str, base_url: str, as_of_month: date) -> Optional[str]:
    """Pick the disclosure-file href referencing as_of_month from a page."""
    pats = _month_patterns(as_of_month)
    for href in re.findall(r'href=["\']([^"\']+)["\']', html, re.I):
        if not _DISCLOSURE_EXT_RE.search(href):
            continue
        low = href.lower()
        if any(re.search(p, low) for p in pats):
            return urljoin(base_url, href.replace(" ", "%20"))
    return None


async def _discover_disclosure_url(
    http: aiohttp.ClientSession, amc_id: str, as_of_month: date,
) -> Optional[str]:
    slug = AK_SLUG.get(amc_id)
    if not slug:
        return None
    page = f"{_AK_BASE}/{slug}/Monthly-Portfolio-Disclosures"
    try:
        async with http.get(page, allow_redirects=True) as resp:
            if resp.status != 200:
                logger.info("advisorkhoj[%s]: page %s → HTTP %d", amc_id, page, resp.status)
                return None
            html = await resp.text()
    except aiohttp.ClientError as e:
        logger.info("advisorkhoj[%s]: page fetch error: %s: %s", amc_id, type(e).__name__, e)
        return None
    url = _file_url_for_month(html, page, as_of_month)
    if not url:
        logger.info("advisorkhoj[%s]: no disclosure file for %s on the page",
                    amc_id, as_of_month.isoformat())
    return url


def _load_sheets(data: bytes) -> list[tuple[str, list]]:
    """Return [(sheet_name, rows)] from an .xlsx (openpyxl) or legacy .xls (xlrd)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return [(name, [list(r) for r in wb[name].iter_rows(values_only=True)])
                for name in wb.sheetnames]
    except Exception:                                        # noqa: BLE001 — try .xls next
        pass
    if not _XLRD:
        return []
    try:
        wb = _xlrd.open_workbook(file_contents=data)
    except Exception:                                        # noqa: BLE001
        return []
    return [(sh.name, [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)])
            for sh in wb.sheets()]


_SCHEME_PREFIX_RE = re.compile(r"^\s*(?:portfolio\s+(?:statement\s+)?of|scheme\s*[:\-])\s*", re.I)
_SCHEME_SUFFIX_RE = re.compile(r"\s+as\s+(?:on|at|of)\b.*$", re.I)        # drop 'as on 31-May-2026'
_SCHEME_REJECT_RE = re.compile(
    r"portfolio statement|monthly portfolio|disclos|provisional|mutual fund$", re.I)


def _scheme_from_sheet(rows: list) -> Optional[str]:
    """Pull the fund name from a sheet's title rows — handles 'Portfolio of
    Kotak X Fund as on …', 'SCHEME: UTI X', or a plain 'AMC … Fund' cell."""
    for r in rows[:5]:
        for c in r:
            if not c:
                continue
            raw = re.sub(r"\s+", " ", str(c)).strip()
            had_prefix = bool(_SCHEME_PREFIX_RE.match(raw))
            s = _SCHEME_SUFFIX_RE.sub("", _SCHEME_PREFIX_RE.sub("", raw)).strip()
            if not (8 < len(s) < 90) or _SCHEME_REJECT_RE.search(s):
                continue
            if had_prefix or re.search(r"\b(fund|plan|etf)\b", s, re.I):
                return s
    return None


def _find_cols(rows: list) -> Optional[dict]:
    """Locate the header row + columns. ISIN and name columns are confirmed
    against the data so merged/offset name columns (Kotak) are handled."""
    isin_counts: dict[int, int] = {}
    for r in rows:
        for j, c in enumerate(r):
            if c and _ISIN_RE.match(str(c).strip().replace("\t", "")):
                isin_counts[j] = isin_counts.get(j, 0) + 1
    if not isin_counts:
        return None
    isin_col = max(isin_counts, key=lambda k: isin_counts[k])
    if isin_counts[isin_col] < 2:
        return None

    header_idx = None
    header: list[str] = []
    for i, r in enumerate(rows[:40]):
        low = [str(c or "").strip().lower() for c in r]
        if any(any(k in c for k in _H_NAME) for c in low):
            header_idx, header = i, low
            break
    if header_idx is None:
        return None

    def hcol(*keys: str) -> Optional[int]:
        for j, c in enumerate(header):
            if any(k in c for k in keys):
                return j
        return None

    data_rows = [r for r in rows[header_idx + 1:]
                 if isin_col < len(r) and r[isin_col]
                 and _ISIN_RE.match(str(r[isin_col]).strip().replace("\t", ""))]

    def textiness(col: int) -> int:
        n = 0
        for r in data_rows[:30]:
            v = r[col] if col < len(r) else None
            if isinstance(v, str) and any(ch.isalpha() for ch in v) and not _ISIN_RE.match(v.strip()):
                n += 1
        return n

    name_col = hcol(*_H_NAME)
    if name_col is None or textiness(name_col) < max(1, len(data_rows[:30]) // 3):
        # Fall back to the most text-bearing column left of ISIN (merged-cell layouts).
        cands = [c for c in range(len(header)) if c != isin_col]
        left = [c for c in cands if c < isin_col] or cands
        if left:
            name_col = max(left, key=textiness)
    return {
        "header_idx": header_idx, "isin": isin_col, "name": name_col,
        "wt": hcol(*_H_WT), "mv": hcol(*_H_MV),
        "ind": hcol("industry", "rating", "sector"), "qty": hcol("quantity", "qty"),
    }


def _parse_sheet_holdings(rows: list, scheme: str, as_of_str: str,
                          source_url: str, source_tag: str) -> list[dict]:
    """ISIN-keyed holdings from one scheme sheet; weight normalised to 0–100."""
    cols = _find_cols(rows)
    if not cols or cols["name"] is None:
        return []

    def cell(r: list, key: str):
        j = cols[key]
        return r[j] if j is not None and j < len(r) else None

    out: list[dict] = []
    for r in rows[cols["header_idx"] + 1:]:
        iv = cell(r, "isin")
        isin = str(iv).strip().replace("\t", "") if iv else ""
        if not _ISIN_RE.match(isin):
            continue                                   # drops aggregates/cash/TREPS/derivatives
        name = str(cell(r, "name") or "").strip()
        if not name:
            continue
        mv_lakh = _to_float(cell(r, "mv"))
        sector = cell(r, "ind")
        out.append({
            "scheme_code": scheme,                     # resolved to AMFI code by the adapter
            "as_of_month": as_of_str,
            "security_isin": isin,
            "security_name": name,
            "instrument_type": None,
            "sector": str(sector).strip() if sector else None,
            "rating": None,
            "quantity": _to_float(cell(r, "qty")),
            "market_value_inr": (mv_lakh * 100000.0) if mv_lakh is not None else None,  # ₹lakh → ₹
            "weight_pct": _to_float(cell(r, "wt")),
            "source": source_tag,
            "source_url": source_url,
        })

    weights = [h["weight_pct"] for h in out if h["weight_pct"] is not None]
    if weights and sum(weights) < 2.5:                 # fraction (sum≈1) → percent
        for h in out:
            if h["weight_pct"] is not None:
                h["weight_pct"] = round(h["weight_pct"] * 100, 4)
    return out


# Sheets that are not a scheme portfolio (index/summary/disclaimer riders).
_SKIP_SHEET_RE = re.compile(r"^(index|cover|content|disclaimer|risk|notes?|summary)\b", re.I)


def parse_disclosure(data: bytes, ext: str, as_of_month: date,
                     source_url: str, source_tag: str,
                     fallback_scheme: Optional[str] = None) -> list[dict]:
    """Parse a monthly disclosure into holding rows (scheme_code = fund NAME).

    Handles a ZIP (one workbook per fund, or a consolidated workbook inside),
    a single per-fund workbook, and a single CONSOLIDATED workbook (one sheet
    per scheme). .xlsx and legacy .xls both supported.
    """
    as_of_str = date(as_of_month.year, as_of_month.month, 1).isoformat()

    if ext == "zip":
        out: list[dict] = []
        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
        except zipfile.BadZipFile:
            logger.warning("advisorkhoj: not a valid zip (%s)", source_url)
            return out
        for name in zf.namelist():
            if not name.lower().endswith((".xlsx", ".xls")):
                continue
            stem = name.rsplit("/", 1)[-1].rsplit(".", 1)[0].strip()
            member_ext = name.rsplit(".", 1)[-1].lower()
            try:
                out.extend(parse_disclosure(zf.read(name), member_ext, as_of_month,
                                            source_url, source_tag, fallback_scheme=stem))
            except Exception as e:                          # noqa: BLE001
                logger.debug("advisorkhoj: member %r parse skipped: %s", stem, e)
        return out

    sheets = _load_sheets(data)
    if not sheets:
        return []

    # Consolidated workbook (one sheet per scheme) vs single per-fund workbook.
    if len(sheets) > 3:
        out = []
        for name, rows in sheets:
            if _SKIP_SHEET_RE.match(name.strip()):
                continue
            scheme = _scheme_from_sheet(rows) or name.strip()
            out.extend(_parse_sheet_holdings(rows, scheme, as_of_str, source_url, source_tag))
        return out

    # Single fund: the first sheet is the portfolio (a 'Derivative' detail
    # sheet, if any, is skipped). Prefer the filename (cleaner than the
    # in-sheet title), then the sheet title.
    name, rows = sheets[0]
    scheme = fallback_scheme or _scheme_from_sheet(rows) or name.strip()
    return _parse_sheet_holdings(rows, scheme, as_of_str, source_url, source_tag)


def parse_disclosure_zip(zip_bytes: bytes, as_of_month: date,
                         source_url: str, source_tag: str) -> list[dict]:
    """Back-compat wrapper: parse a disclosure ZIP into holding rows."""
    return parse_disclosure(zip_bytes, "zip", as_of_month, source_url, source_tag)


async def advisorkhoj_holdings_adapter(
    amc_id: str, http: aiohttp.ClientSession, as_of_month: date,
) -> list[dict]:
    """Discover → download → parse → resolve the monthly disclosure for amc_id.

    Returns mf_holdings_monthly-shaped rows with scheme_code resolved to
    AMFI codes (one fund's holdings fan out to all its plan variants).
    Returns [] if no disclosure file is found or nothing parses.
    """
    from nidp.shared.storage.pg import get_pool
    from .amc_dispatch import _build_scheme_index, _resolve_scheme_codes

    url = await _discover_disclosure_url(http, amc_id, as_of_month)
    if not url:
        return []
    try:
        async with http.get(url, allow_redirects=True) as resp:
            if resp.status != 200:
                logger.info("advisorkhoj[%s]: file %s → HTTP %d", amc_id, url, resp.status)
                return []
            data = await resp.read()
    except aiohttp.ClientError as e:
        logger.info("advisorkhoj[%s]: download error: %s: %s", amc_id, type(e).__name__, e)
        return []

    source_tag = f"{amc_id.upper()}_DISCLOSURE_ADVISORKHOJ"
    ext = url.split("?")[0].rsplit(".", 1)[-1].lower()
    if ext not in ("zip", "xlsx", "xls"):
        ext = "zip"
    parsed = parse_disclosure(data, ext, as_of_month, url, source_tag)
    if not parsed:
        logger.info("advisorkhoj[%s]: %s parsed 0 holdings", amc_id, url)
        return []

    # Resolve fund name → AMFI scheme_codes and fan out.
    pool = await get_pool()
    async with pool.acquire() as conn:
        db = await conn.fetch(
            "SELECT scheme_code, scheme_name FROM nidp.mf_scheme_master WHERE amc_id = $1",
            amc_id,
        )
    index = _build_scheme_index(
        [{"scheme_code": r["scheme_code"], "scheme_name": r["scheme_name"]} for r in db]
    )

    out: list[dict] = []
    code_cache: dict[str, list[str]] = {}
    funds_resolved: set[str] = set()
    for row in parsed:
        fund = row["scheme_code"]
        codes = code_cache.get(fund)
        if codes is None:
            codes = _resolve_scheme_codes(fund, index)
            code_cache[fund] = codes
        if not codes:
            continue
        funds_resolved.add(fund)
        for code in codes:
            out.append({**row, "scheme_code": str(code)})

    logger.info("advisorkhoj[%s]: %d funds → %d resolved → %d holding rows over %d scheme_codes",
                amc_id, len(code_cache), len(funds_resolved), len(out),
                len({r["scheme_code"] for r in out}))
    return out
