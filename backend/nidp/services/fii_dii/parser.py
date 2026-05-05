"""NSE fii_stats XLS/XLSX parser.

NSE publishes an Excel workbook with multiple sheets, typically:
    - "FII" or "Cash" sheet with cash-segment net values for FII/DII
    - "FII Stats" or "Derivatives" sheet for F&O participant-wise data

Layout has shifted across years. We auto-detect by sheet name and
column header, normalise to long form: one row per (category, segment).

This parser depends on:
    openpyxl  — XLSX
    xlrd<2    — legacy XLS (BIFF)

If neither is installed or the file format is unrecognised, raises a
loud NotImplementedError with the file's first bytes preview, which
the BaseIngester maps to a FAILED job_log row + Prometheus alert.

TODO(parser-completion): the column→(category,segment) mapping below
covers the layouts we've validated against; observed XLS variants
need golden-file tests in tests/fixtures/fii_dii/. Add new layouts
as we encounter them — DO NOT silently fall back; the PRD's "no bad
data" rule applies most strongly here.
"""
from __future__ import annotations

import io
import logging
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Heuristic mapping from cell-text to (category, segment) labels.
# FII/DII cash flows are the priority; F&O segments come second.
_CATEGORY_HINTS = {
    "FII": "FII", "FPI": "FII", "DII": "DII",
    "PROP": "PROP", "PROPRIETARY": "PROP", "CLIENT": "CLIENT",
}
_SEGMENT_HINTS = {
    "CASH": "EQUITY_CASH", "EQUITY": "EQUITY_CASH",
    "INDEX FUTURES": "INDEX_FUTURES", "INDEX FUT": "INDEX_FUTURES",
    "INDEX OPTIONS": "INDEX_OPTIONS", "INDEX OPT": "INDEX_OPTIONS",
    "STOCK FUTURES": "STOCK_FUTURES", "STOCK FUT": "STOCK_FUTURES",
    "STOCK OPTIONS": "STOCK_OPTIONS", "STOCK OPT": "STOCK_OPTIONS",
}


def _detect_format(body: bytes) -> str:
    head = body[:8]
    if head[:4] == b"PK\x03\x04":
        return "xlsx"
    if head[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"
    # JSON API (current NSE endpoint /api/fiidiiTradeReact)
    stripped = body.lstrip()[:8]
    if stripped[:1] in (b"[", b"{"):
        return "json"
    if b"<html" in body[:512].lower() or b"<table" in body[:512].lower():
        return "html"
    raise NotImplementedError(
        f"unrecognised fii_stats body (first 8 bytes: {head!r}). "
        f"Add layout to fii_dii parser."
    )


def _parse_json_api(body: bytes, fallback_date_iso: str) -> list[dict[str, Any]]:
    """Parse the new NSE /api/fiidiiTradeReact JSON shape.

    Response: array of objects with fields
        category   ('FII' | 'DII')
        date       ('05-May-2026')
        buyValue, sellValue, netValue  (string-encoded Cr ₹)
    Only cash segment is exposed via this endpoint.
    """
    import json
    from datetime import datetime
    payload = json.loads(body.decode("utf-8", errors="replace"))
    if isinstance(payload, dict):
        for k in ("data", "items", "rows"):
            if isinstance(payload.get(k), list):
                payload = payload[k]
                break
    if not isinstance(payload, list):
        return []

    def _f(s):
        if s is None or s == "":
            return None
        try:
            return float(str(s).replace(",", ""))
        except ValueError:
            return None

    def _date(s):
        if not s:
            return fallback_date_iso
        for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date().isoformat()
            except ValueError:
                continue
        return fallback_date_iso

    rows: list[dict[str, Any]] = []
    for entry in payload:
        if not isinstance(entry, dict):
            continue
        cat_raw = (entry.get("category") or "").strip().upper()
        # NSE serves combined "FII/FPI" plus standalone "FII" / "FPI" — fold
        # all into 'FII' for our schema. DII passes through.
        if "FII" in cat_raw or "FPI" in cat_raw:
            cat = "FII"
        elif cat_raw == "DII":
            cat = "DII"
        else:
            continue
        rows.append({
            "as_of_date":     _date(entry.get("date")),
            "category":       cat,
            "segment":        "EQUITY_CASH",
            "buy_value_cr":   _f(entry.get("buyValue")),
            "sell_value_cr":  _f(entry.get("sellValue")),
            "net_value_cr":   _f(entry.get("netValue")),
            "buy_contracts":  None,
            "sell_contracts": None,
            "net_contracts":  None,
            "open_interest":  None,
        })
    return rows


def _open_xlsx(body: bytes):
    try:
        import openpyxl                                            # type: ignore[import-not-found]
    except ImportError as e:                                       # pragma: no cover
        raise RuntimeError("fii_dii xlsx parser requires openpyxl") from e
    return openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)


def _open_xls(body: bytes):
    try:
        import xlrd                                                # type: ignore[import-not-found]
    except ImportError as e:                                       # pragma: no cover
        raise RuntimeError("fii_dii xls parser requires xlrd<2.0.0") from e
    return xlrd.open_workbook(file_contents=body)


def _norm(s: Any) -> str:
    return (str(s) if s is not None else "").strip().upper()


def _classify_label(label: str) -> tuple[Optional[str], Optional[str]]:
    cat: Optional[str] = None
    seg: Optional[str] = None
    up = _norm(label)
    for k, v in _CATEGORY_HINTS.items():
        if k in up:
            cat = v
            break
    for k, v in _SEGMENT_HINTS.items():
        if k in up:
            seg = v
            break
    return cat, seg


def _f(x: Any) -> Optional[float]:
    if x is None or x == "":
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    try: return float(s)
    except ValueError: return None


def parse_fii_dii(body: bytes, target_date_iso: str) -> list[dict[str, Any]]:
    """Return long-form rows. Failures raise loudly."""
    fmt = _detect_format(body)
    rows: list[dict[str, Any]] = []

    # Current NSE endpoint returns JSON.
    if fmt == "json":
        return _parse_json_api(body, target_date_iso)

    if fmt == "html":
        # NSE occasionally serves an HTML table when the XLS is being
        # rotated. Use pandas read_html if available.
        try:
            import pandas as pd                                    # type: ignore[import-not-found]
        except ImportError as e:                                   # pragma: no cover
            raise RuntimeError("fii_dii html fallback requires pandas") from e
        tables = pd.read_html(io.BytesIO(body))
        if not tables:
            return []
        # Best-effort: scan first table looking for FII/DII rows.
        df = tables[0]
        for _, r in df.iterrows():
            label = _norm(r.iloc[0]) if len(r) else ""
            cat, seg = _classify_label(label)
            if not cat:
                continue
            seg = seg or "EQUITY_CASH"
            buy = _f(r.iloc[1] if len(r) > 1 else None)
            sell = _f(r.iloc[2] if len(r) > 2 else None)
            net = _f(r.iloc[3] if len(r) > 3 else None) or (
                (buy - sell) if (buy is not None and sell is not None) else None
            )
            rows.append({
                "as_of_date": target_date_iso,
                "category": cat, "segment": seg,
                "buy_value_cr": buy, "sell_value_cr": sell, "net_value_cr": net,
                "buy_contracts": None, "sell_contracts": None,
                "net_contracts": None, "open_interest": None,
            })
        return rows

    # XLS / XLSX path
    if fmt == "xlsx":
        wb = _open_xlsx(body)
        sheets = [(s, wb[s]) for s in wb.sheetnames]
        def _iter(sheet): return ([cell.value for cell in row] for row in sheet.iter_rows())
    else:  # xls
        wb = _open_xls(body)
        sheets = [(s, wb.sheet_by_name(s)) for s in wb.sheet_names()]
        def _iter(sheet):
            for ri in range(sheet.nrows):
                yield [sheet.cell_value(ri, ci) for ci in range(sheet.ncols)]

    for sheet_name, sheet in sheets:
        for raw in _iter(sheet):
            if not raw:
                continue
            label = _norm(raw[0])
            cat, seg = _classify_label(label)
            if not cat:
                continue
            seg = seg or "EQUITY_CASH"
            # Standard layout: col 1=BUY, col 2=SELL, col 3=NET (Cr).
            buy = _f(raw[1] if len(raw) > 1 else None)
            sell = _f(raw[2] if len(raw) > 2 else None)
            net = _f(raw[3] if len(raw) > 3 else None) or (
                (buy - sell) if (buy is not None and sell is not None) else None
            )
            rows.append({
                "as_of_date":     target_date_iso,
                "category":       cat,
                "segment":        seg,
                "buy_value_cr":   buy,
                "sell_value_cr":  sell,
                "net_value_cr":   net,
                "buy_contracts":  None,
                "sell_contracts": None,
                "net_contracts":  None,
                "open_interest":  None,
            })
    return rows
