"""AMFI central disclosure fetchers: TER and risk-o-meter.

AMFI publishes both disclosures for *all* schemes in a single Excel per month.
Fetching once and caching in-process avoids redundant downloads when multiple
per-AMC adapters run in the same service call.

Typical data shape:
  TER Excel:         Scheme Code | Scheme Name | TER (Regular) % | TER (Direct) %
  Risk-o-meter Excel: Scheme Code | Scheme Name | Risk Category
"""
from __future__ import annotations

import io
import logging
import re
from typing import Optional

import aiohttp
import openpyxl
from bs4 import BeautifulSoup

from nidp.shared.config import AMFI_WWW

logger = logging.getLogger(__name__)

TER_PAGE_URL  = f"{AMFI_WWW}/research-information/other-data/ter"
RISK_PAGE_URL = f"{AMFI_WWW}/research-information/other-data/risk-o-meter"

# Fallback candidates for when AMFI reshuffles their site (every 1-2y).
# Drift-check job pings these daily and alerts on red.
_TER_PAGE_CANDIDATES: tuple[str, ...] = (
    TER_PAGE_URL,
    f"{AMFI_WWW}/other-data/ter",
    f"{AMFI_WWW}/research-information/ter",
    f"{AMFI_WWW}/other-data/expense-ratio",
)
_RISK_PAGE_CANDIDATES: tuple[str, ...] = (
    RISK_PAGE_URL,
    f"{AMFI_WWW}/other-data/risk-o-meter",
    f"{AMFI_WWW}/research-information/risk-o-meter",
    f"{AMFI_WWW}/other-data/riskometer",
)

_TER_CACHE:  Optional[dict[str, tuple[Optional[float], Optional[float]]]] = None
_RISK_CACHE: Optional[dict[str, str]] = None


async def _fetch_first_xlsx_url(
    http: aiohttp.ClientSession, candidates: tuple[str, ...], label: str,
) -> Optional[str]:
    """Walk candidate landing pages in order; return first xlsx link found.

    None when no candidate yields one — caller logs structured warning.
    """
    for page in candidates:
        try:
            async with http.get(page, allow_redirects=True) as resp:
                if resp.status != 200:
                    logger.debug("amfi_central[%s]: %s → status=%d", label, page, resp.status)
                    continue
                html = await resp.text(errors="replace")
        except Exception as e:  # noqa: BLE001
            logger.debug("amfi_central[%s]: %s → %s: %s", label, page, type(e).__name__, e)
            continue
        url = _latest_xlsx_link(html)
        if url:
            logger.info("amfi_central[%s]: discovered xlsx via %s → %s", label, page, url)
            return url
        logger.debug("amfi_central[%s]: %s loaded but no xlsx link found", label, page)
    return None


def _latest_xlsx_link(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, "lxml")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(r"\.(xlsx|xls)($|\?)", href, re.IGNORECASE):
            if not href.startswith("http"):
                href = AMFI_WWW + "/" + href.lstrip("/")
            links.append(href)
    return links[-1] if links else None


def _col_idx(headers: list[str], *fragments: str) -> Optional[int]:
    for frag in fragments:
        for i, h in enumerate(headers):
            if frag in h:
                return i
    return None


def _to_float(v: object) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None


def _parse_ter_xlsx(data: bytes) -> dict[str, tuple[Optional[float], Optional[float]]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return {}

    header_row = 0
    for i, row in enumerate(raw):
        cells = [str(c or "").lower() for c in row]
        if any("scheme" in c or "code" in c for c in cells):
            header_row = i
            break

    headers = [str(c or "").strip().lower() for c in raw[header_row]]
    code_col  = _col_idx(headers, "scheme code", "amfi code", "schemecode")
    ter_r_col = _col_idx(headers, "ter (regular", "regular plan", "ter%", "ter (r")
    ter_d_col = _col_idx(headers, "ter (direct", "direct plan", "ter (d")

    if code_col is None:
        logger.warning("amfi_central: TER xlsx — scheme-code column not found; headers=%s", headers)
        return {}

    result: dict[str, tuple[Optional[float], Optional[float]]] = {}
    for row in raw[header_row + 1:]:
        if not row or row[code_col] is None:
            continue
        code = str(row[code_col]).strip()
        if not code or not code[0].isdigit():
            continue
        def _f(col: Optional[int]) -> Optional[float]:
            if col is None or col >= len(row):
                return None
            return _to_float(row[col])
        result[code] = (_f(ter_r_col), _f(ter_d_col))
    return result


def _parse_risk_xlsx(data: bytes) -> dict[str, str]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return {}

    header_row = 0
    for i, row in enumerate(raw):
        cells = [str(c or "").lower() for c in row]
        if any("scheme" in c or "code" in c for c in cells):
            header_row = i
            break

    headers = [str(c or "").strip().lower() for c in raw[header_row]]
    code_col = _col_idx(headers, "scheme code", "amfi code", "schemecode")
    risk_col = _col_idx(headers, "risk category", "risk-o-meter", "riskometer", "risk level", "risk grade")

    if code_col is None or risk_col is None:
        logger.warning("amfi_central: risk xlsx — columns not found; headers=%s", headers)
        return {}

    result: dict[str, str] = {}
    for row in raw[header_row + 1:]:
        if not row or row[code_col] is None:
            continue
        code = str(row[code_col]).strip()
        risk = str(row[risk_col] or "").strip()
        if code and code[0].isdigit() and risk:
            result[code] = risk
    return result


async def fetch_ter_all(
    http: aiohttp.ClientSession,
) -> dict[str, tuple[Optional[float], Optional[float]]]:
    """Return {scheme_code: (ter_regular_pct, ter_direct_pct)} for all AMFI schemes.
    Module-level cache: fetched once per process per run.
    """
    global _TER_CACHE
    if _TER_CACHE is not None:
        return _TER_CACHE

    try:
        url = await _fetch_first_xlsx_url(http, _TER_PAGE_CANDIDATES, "ter")
        if not url:
            logger.warning(
                "amfi_central: no TER xlsx link on any of %d candidates "
                "(first: %s) — AMFI may have reorganised the page",
                len(_TER_PAGE_CANDIDATES), _TER_PAGE_CANDIDATES[0],
            )
            _TER_CACHE = {}
            return _TER_CACHE
        async with http.get(url) as resp:
            resp.raise_for_status()
            data = await resp.read()
        _TER_CACHE = _parse_ter_xlsx(data)
        logger.info("amfi_central: TER loaded %d schemes from %s", len(_TER_CACHE), url)
    except Exception as e:  # noqa: BLE001
        logger.warning("amfi_central: TER fetch error: %s: %s", type(e).__name__, e)
        _TER_CACHE = {}
    return _TER_CACHE


async def fetch_risk_all(http: aiohttp.ClientSession) -> dict[str, str]:
    """Return {scheme_code: risk_label} for all AMFI schemes."""
    global _RISK_CACHE
    if _RISK_CACHE is not None:
        return _RISK_CACHE

    try:
        url = await _fetch_first_xlsx_url(http, _RISK_PAGE_CANDIDATES, "risk")
        if not url:
            logger.warning(
                "amfi_central: no risk xlsx link on any of %d candidates "
                "(first: %s) — AMFI may have reorganised the page",
                len(_RISK_PAGE_CANDIDATES), _RISK_PAGE_CANDIDATES[0],
            )
            _RISK_CACHE = {}
            return _RISK_CACHE
        async with http.get(url) as resp:
            resp.raise_for_status()
            data = await resp.read()
        _RISK_CACHE = _parse_risk_xlsx(data)
        logger.info("amfi_central: risk-o-meter loaded %d schemes from %s", len(_RISK_CACHE), url)
    except Exception as e:  # noqa: BLE001
        logger.warning("amfi_central: risk fetch error: %s: %s", type(e).__name__, e)
        _RISK_CACHE = {}
    return _RISK_CACHE


def clear_cache() -> None:
    """Reset in-process caches (testing / forced refresh)."""
    global _TER_CACHE, _RISK_CACHE
    _TER_CACHE = None
    _RISK_CACHE = None
