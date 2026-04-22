"""Tests for /api/admin/v3-master-funds (JSON catalogue + Excel export).

Verifies shape, required fields, formula reference, and Excel integrity.
"""
import io
import os
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
ADMIN_TOKEN = "370eff71-fda1-46d8-b506-b81b894d634f"  # priyankamantri@gmail.com (admin)


@pytest.fixture(scope="module")
def api():
    if not BASE_URL:
        pytest.skip("REACT_APP_BACKEND_URL not set")
    s = requests.Session()
    s.cookies.set("session_token", ADMIN_TOKEN)
    return s


@pytest.fixture(scope="module")
def payload(api):
    r = api.get(f"{BASE_URL}/api/admin/v3-master-funds?limit=10", timeout=60)
    assert r.status_code == 200, f"{r.status_code}: {r.text[:200]}"
    return r.json()


# ── JSON contract ────────────────────────────────────────────────────────
def test_payload_has_top_level_keys(payload):
    for k in ("generated_at", "totals", "formula_reference", "funds"):
        assert k in payload, f"missing {k}"


def test_totals_counts_consistent(payload):
    t = payload["totals"]
    assert t["total"] == len(payload["funds"])
    assert t["scored"] <= t["total"]
    assert sum(t["by_source"].values()) == t["total"]
    assert sum(t["by_category"].values()) == t["total"]


def test_formula_reference_has_all_scores(payload):
    fr = payload["formula_reference"]
    for score in ("quality", "health", "exit", "add", "switch"):
        assert score in fr, f"formula_reference missing {score}"
    # Quality + Health must expose weights
    for s in ("quality", "health"):
        assert fr[s].get("weights"), f"{s} missing weights"
        assert fr[s].get("component_logic"), f"{s} missing component_logic"
        assert fr[s].get("primitives_used"), f"{s} missing primitives_used"


def test_each_fund_has_scores_block(payload):
    for f in payload["funds"]:
        assert "scores" in f
        assert {"quality", "health", "exit", "add"}.issubset(f["scores"])


def test_each_fund_has_primitives_block(payload):
    expected = {
        "aum_cr", "fund_age_years", "expense_ratio_direct",
        "manager_tenure_years", "top10_concentration_pct",
        "consistency_score", "max_drawdown_pct", "turnover_ratio",
        "sharpe", "sortino",
    }
    for f in payload["funds"]:
        assert expected.issubset(f["primitives"]), (
            f"fund {f['scheme_name']} missing primitives: "
            f"{expected - set(f['primitives'])}"
        )


def test_each_fund_has_quality_and_health_breakdown(payload):
    for f in payload["funds"]:
        qb = f.get("quality_breakdown") or {}
        hb = f.get("health_breakdown") or {}
        assert isinstance(qb.get("components"), dict), "quality_breakdown.components missing"
        assert isinstance(hb.get("components"), dict), "health_breakdown.components missing"
        assert isinstance(qb.get("effective_weights"), dict), "quality_breakdown.effective_weights missing"
        assert isinstance(hb.get("effective_weights"), dict), "health_breakdown.effective_weights missing"
        assert isinstance(qb.get("missing_primitives"), list), "quality_breakdown.missing_primitives missing"


def test_quality_components_are_expected_set(payload):
    expected = {"performance", "risk_adjusted", "consistency", "drawdown", "cost", "aum_age"}
    for f in payload["funds"]:
        got = set(f["quality_breakdown"]["components"].keys())
        assert got == expected, f"quality components mismatch for {f['scheme_name']}: {got}"


def test_health_components_are_expected_set(payload):
    expected = {"manager_tenure", "aum_stability", "turnover", "concentration",
                "downside_protection", "expense_trend"}
    for f in payload["funds"]:
        got = set(f["health_breakdown"]["components"].keys())
        assert got == expected, f"health components mismatch for {f['scheme_name']}: {got}"


def test_effective_weights_sum_to_100_when_all_present(payload):
    """When no primitive is missing, effective weights should sum to 100 ± 0.1."""
    for f in payload["funds"]:
        for kind in ("quality_breakdown", "health_breakdown"):
            b = f[kind]
            if not b.get("missing_primitives"):
                total = sum(b["effective_weights"].values())
                assert abs(total - 100) < 0.2, (
                    f"{kind} effective weights don't sum to 100 for {f['scheme_name']}: {total}"
                )


def test_source_is_groww_or_tickertape(payload):
    for f in payload["funds"]:
        assert f["source"] in ("groww", "tickertape")


def test_plan_type_is_derived(payload):
    for f in payload["funds"]:
        assert f["plan_type"] in ("direct", "regular")


def test_sorted_by_quality_desc(payload):
    scores = [(f["scores"].get("quality") or -1) for f in payload["funds"]]
    for a, b in zip(scores, scores[1:]):
        assert a >= b, f"not sorted by quality desc: {a} < {b}"


# ── Excel export ─────────────────────────────────────────────────────────
def test_excel_export_returns_xlsx(api):
    r = api.get(f"{BASE_URL}/api/admin/v3-master-funds/export.xlsx", timeout=120)
    assert r.status_code == 200
    assert r.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(r.content) > 5000, "xlsx suspiciously small"


def test_excel_has_all_expected_sheets(api):
    import openpyxl
    r = api.get(f"{BASE_URL}/api/admin/v3-master-funds/export.xlsx", timeout=120)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    sheets = wb.sheetnames
    for expected in ("Summary", "Primitives", "Quality breakdown",
                     "Health breakdown", "Formula reference"):
        assert expected in sheets, f"missing sheet: {expected} (got {sheets})"


def test_excel_summary_sheet_has_rows(api):
    import openpyxl
    r = api.get(f"{BASE_URL}/api/admin/v3-master-funds/export.xlsx", timeout=120)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb["Summary"]
    # Headers on row 1, funds start row 2
    assert ws.max_row >= 2, "no data rows in Summary"
    # Headers should include score columns
    headers = [c.value for c in ws[1]]
    for h in ("Scheme name", "Quality /100", "Health /100", "AUM (₹ Cr)"):
        assert h in headers, f"Summary missing header {h}"


def test_excel_primitives_sheet_width(api):
    import openpyxl
    r = api.get(f"{BASE_URL}/api/admin/v3-master-funds/export.xlsx", timeout=120)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb["Primitives"]
    headers = [c.value for c in ws[1]]
    for h in ("aum_cr", "fund_age_years", "sharpe", "top10_concentration_pct"):
        assert h in headers, f"Primitives missing column {h}"


# ── Auth ─────────────────────────────────────────────────────────────────
def test_non_admin_cannot_access_json():
    if not BASE_URL:
        pytest.skip("REACT_APP_BACKEND_URL not set")
    r = requests.get(f"{BASE_URL}/api/admin/v3-master-funds", timeout=20)
    assert r.status_code in (401, 403)


def test_non_admin_cannot_export():
    if not BASE_URL:
        pytest.skip("REACT_APP_BACKEND_URL not set")
    r = requests.get(f"{BASE_URL}/api/admin/v3-master-funds/export.xlsx", timeout=20)
    assert r.status_code in (401, 403)
