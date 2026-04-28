"""Portfolio holdings export — CSV / XLSX.

Two entry points:

  • GET /api/portfolio/holdings/export?format=csv|xlsx
        Self-only: exports the logged-in user's (or impersonated profile's)
        holdings via the standard `get_current_user` impersonation contract.
        Honours an active `active_profile_id` automatically.

  • GET /api/mfd/profiles/{profile_id}/holdings/export?format=csv|xlsx
        Advisor-scoped: workspace-checked. The MFD must own the workspace
        the profile lives in. Used by the export icon on each row of the
        advisor dashboard's client list.

Same response shape from both: a streamed CSV / XLSX file with the
client's name + statement date in the filename.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from deps import db, get_current_user

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")


# ── Column spec — drives both CSV header AND XLSX header row ─────────────
EXPORT_COLUMNS: List[Dict[str, str]] = [
    {"key": "name",          "label": "Name",          "kind": "text"},
    {"key": "ticker",        "label": "ISIN/Ticker",   "kind": "text"},
    {"key": "asset_type",    "label": "Asset type",    "kind": "text"},
    {"key": "sector",        "label": "Sector / Cat.", "kind": "text"},
    {"key": "amc",           "label": "AMC",           "kind": "text"},
    {"key": "plan",          "label": "Plan",          "kind": "text"},
    {"key": "option",        "label": "Option",        "kind": "text"},
    {"key": "folio",         "label": "Folio",         "kind": "text"},
    {"key": "quantity",      "label": "Units / Qty",   "kind": "number"},
    {"key": "buy_price",     "label": "Avg cost",      "kind": "number"},
    {"key": "current_price", "label": "Current price", "kind": "number"},
    {"key": "current_value", "label": "Current value", "kind": "number"},  # computed
    {"key": "invested",      "label": "Invested",      "kind": "number"},  # computed
    {"key": "gain_inr",      "label": "P&L (₹)",       "kind": "number"},  # computed
    {"key": "gain_pct",      "label": "P&L (%)",       "kind": "number"},  # computed
    {"key": "source",        "label": "Source",        "kind": "text"},
    {"key": "parsed_by",     "label": "Parsed by",     "kind": "text"},
    {"key": "added_at",      "label": "Added on",      "kind": "text"},
]


def _safe_filename(s: str) -> str:
    """Make `s` safe to embed in a Content-Disposition filename."""
    return re.sub(r"[^a-zA-Z0-9_-]+", "_", s or "").strip("_") or "portfolio"


def _enrich_row(h: Dict[str, Any]) -> Dict[str, Any]:
    """Add computed columns the engine doesn't store directly."""
    qty = float(h.get("quantity") or 0)
    buy = float(h.get("buy_price") or 0)
    cur = float(h.get("current_price") or 0)
    invested = round(qty * buy, 2) if qty and buy else 0
    cur_val = round(qty * cur, 2) if qty and cur else 0
    gain = round(cur_val - invested, 2)
    gain_pct = round((gain / invested) * 100, 2) if invested else 0
    return {
        **h,
        "current_value": cur_val,
        "invested":      invested,
        "gain_inr":      gain,
        "gain_pct":      gain_pct,
    }


async def _load_holdings(user_id: str) -> List[Dict[str, Any]]:
    docs = await db.holdings.find({"user_id": user_id}, {"_id": 0}).sort("name", 1).to_list(2000)
    return [_enrich_row(d) for d in docs]


# ── CSV builder ──────────────────────────────────────────────────────────
def _build_csv(rows: List[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow([c["label"] for c in EXPORT_COLUMNS])
    for r in rows:
        w.writerow([_format_cell(r.get(c["key"]), c["kind"]) for c in EXPORT_COLUMNS])
    return buf.getvalue().encode("utf-8-sig")  # BOM helps Excel open as UTF-8


def _format_cell(v: Any, kind: str) -> str:
    if v is None or v == "":
        return ""
    if kind == "number":
        try:
            return f"{float(v):.2f}"
        except (TypeError, ValueError):
            return str(v)
    return str(v)


# ── XLSX builder ─────────────────────────────────────────────────────────
def _build_xlsx(rows: List[Dict[str, Any]], owner_name: str = "Portfolio") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"[:31]

    # Header — bold white on indigo
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # Title row
    ws.cell(row=1, column=1, value=f"{owner_name} — Holdings export").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = Font(italic=True, color="6B7280", size=10)
    ws.cell(row=2, column=4, value=f"{len(rows)} rows").font = Font(italic=True, color="6B7280", size=10)

    # Header row at row 4
    for col_idx, col in enumerate(EXPORT_COLUMNS, start=1):
        c = ws.cell(row=4, column=col_idx, value=col["label"])
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Data rows
    for r_idx, row in enumerate(rows, start=5):
        for c_idx, col in enumerate(EXPORT_COLUMNS, start=1):
            val = row.get(col["key"])
            if col["kind"] == "number":
                try:
                    val = float(val) if val not in (None, "") else None
                except (TypeError, ValueError):
                    val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = right if col["kind"] == "number" else left
            if col["kind"] == "number" and col["key"] in ("current_value", "invested", "gain_inr"):
                cell.number_format = "#,##0.00"
            elif col["kind"] == "number" and col["key"] == "gain_pct":
                cell.number_format = "0.00\"%\""
            elif col["kind"] == "number":
                cell.number_format = "#,##0.0000"

    # Totals row
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, col in enumerate(EXPORT_COLUMNS, start=1):
        if col["key"] in ("current_value", "invested", "gain_inr"):
            letter = get_column_letter(col_idx)
            ws.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({letter}5:{letter}{total_row - 1})",
            ).font = Font(bold=True)
            ws.cell(row=total_row, column=col_idx).number_format = "#,##0.00"

    # Auto-size columns (cap to 38 chars to keep things readable)
    for col_idx, col in enumerate(EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(col["label"])
        for r in rows[:200]:  # sample first 200 to estimate width
            v = str(r.get(col["key"]) or "")
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)

    # Freeze header
    ws.freeze_panes = "A5"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Public response helper ───────────────────────────────────────────────
def _stream_export(rows: List[Dict[str, Any]], fmt: str, owner_name: str) -> StreamingResponse:
    fmt = (fmt or "csv").lower()
    base = _safe_filename(owner_name)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    if fmt == "xlsx":
        body = _build_xlsx(rows, owner_name=owner_name)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{base}_holdings_{stamp}.xlsx"
    elif fmt == "csv":
        body = _build_csv(rows)
        media = "text/csv; charset=utf-8"
        filename = f"{base}_holdings_{stamp}.csv"
    else:
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")
    return StreamingResponse(
        io.BytesIO(body),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Routes ───────────────────────────────────────────────────────────────
@router.get("/portfolio/holdings/export")
async def export_self_holdings(request: Request, format: str = "csv"):
    """Export holdings for the logged-in user (or impersonated profile when
    the MFD has activated a client). Returns 404 when there are zero
    holdings — the UI renders a tooltip ("Nothing to export") instead of
    downloading an empty file in that case."""
    user = await get_current_user(request)
    rows = await _load_holdings(user["user_id"])
    if not rows:
        raise HTTPException(status_code=404, detail="No holdings to export")
    owner = user.get("name") or user.get("email") or "portfolio"
    return _stream_export(rows, format, owner)


@router.get("/mfd/profiles/{profile_id}/holdings/export")
async def export_profile_holdings(profile_id: str, request: Request, format: str = "csv"):
    """Advisor-scoped export. Workspace ownership is verified:
      1. Resolve the profile → its workspace.
      2. Confirm the workspace is owned by the logged-in user.
      3. Stream the profile's shadow-user holdings.
    """
    user = await get_current_user(request)
    real_uid = user.get("_session_user_id") or user["user_id"]

    prof = await db.profiles.find_one(
        {"profile_id": profile_id},
        {"_id": 0, "shadow_user_id": 1, "workspace_id": 1, "name": 1},
    )
    if not prof or not prof.get("shadow_user_id"):
        raise HTTPException(status_code=404, detail="Profile not found")

    ws = await db.workspaces.find_one(
        {"workspace_id": prof["workspace_id"], "owner_user_id": real_uid},
        {"_id": 0, "workspace_id": 1},
    )
    if not ws:
        raise HTTPException(status_code=403, detail="You don't own this profile's workspace")

    rows = await _load_holdings(prof["shadow_user_id"])
    if not rows:
        raise HTTPException(status_code=404, detail="No holdings to export for this client")
    return _stream_export(rows, format, prof.get("name") or "Client")
