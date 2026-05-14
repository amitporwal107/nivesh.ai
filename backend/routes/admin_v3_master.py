"""Admin V3 Master-Funds dashboard endpoints.

Serves a catalogue view of every mutual fund that has V3 primitives
ingested (Groww and Tickertape), including every raw primitive used by the
V3 engine and the effective weights/contributions of each sub-score.

Endpoints:
  GET  /api/admin/v3-master-funds          → JSON list (with metadata &
                                             formula reference)
  GET  /api/admin/v3-master-funds/export.xlsx
                                           → Multi-sheet workbook
                                             (Summary · Primitives ·
                                              Quality · Health · Formula)
"""
from __future__ import annotations
import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from deps import require_admin
from services import pg_client, v3_scoring

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api")


# ── Static formula reference (used by UI & Excel) ────────────────────────
FORMULA_REFERENCE: Dict[str, Any] = {
    "quality": {
        "formula": "Σ(component × weight) / Σ(available weights) × 10",
        "scale": "0-100",
        "weights": {
            "performance": 25,
            "risk_adjusted": 20,
            "consistency": 20,
            "drawdown": 15,
            "cost": 10,
            "aum_age": 10,
        },
        "component_logic": {
            "performance": "Weighted excess return over category (1Y:20%, 3Y:40%, 5Y:40%). "
                           "Score = 5 + excess_pct/2, clamped [0,10].",
            "risk_adjusted": "Sharpe & Sortino × 6.67, clamped [0,10]. "
                             "Mean of both when present (0 → 0, 1.5 → 10).",
            "consistency": "consistency_score × 10 (fraction of rolling 12-m windows "
                           "beating category). 0.40 → 4.0, 0.80 → 8.0.",
            "drawdown": "Lower drawdown is better. 10 − |dd_pct|/3, clamped [0,10]. "
                        "-15% DD → 5.0.",
            "cost": "Lower expense ratio = higher cost score. 10 − er_pct × 5, clamped [0,10].",
            "aum_age": "AUM > ₹500Cr AND age > 3 yrs required. Mean of log(aum)/log(1L) and "
                       "min(age/10, 1) × 10.",
        },
        "primitives_used": [
            "ret_1y", "ret_3y", "ret_5y",
            "cat_avg_1y", "cat_avg_3y", "cat_avg_5y",
            "sharpe", "sortino",
            "consistency_score",
            "max_drawdown_pct",
            "expense_ratio_direct | expense_ratio",
            "aum_cr", "fund_age_years",
        ],
    },
    "health": {
        "formula": "Σ(component × weight) / Σ(available weights) × 10",
        "scale": "0-100",
        "weights": {
            "manager_tenure": 25,
            "aum_stability": 20,
            "turnover": 15,
            "concentration": 15,
            "downside_protection": 15,
            "expense_trend": 10,
        },
        "component_logic": {
            "manager_tenure": "min(tenure/10, 1) × 10. 0.5y → 0.5, 10y → 10.",
            "aum_stability": "aum_trend_score × 10 (rolling 12-m AUM stability ratio).",
            "turnover": "Lower turnover = higher score. 10 − turnover_pct/20, clamped [0,10]. "
                        "30% → 8.5, 100% → 5.0, 200% → 0.",
            "concentration": "Top-10 concentration. Score = 10 − (top10_pct − 40)/4 "
                             "for equity, clamped [0,10]. <40% → 10, 80% → 0.",
            "downside_protection": "Lower downside-capture = better. "
                                   "10 − (capture_pct − 50)/5, clamped [0,10].",
            "expense_trend": "ER delta over 3 yrs. 10 − max(0, delta × 20), "
                             "clamped [0,10]. Flat ER → 10, +0.5% ER → 0.",
        },
        "primitives_used": [
            "manager_tenure_years",
            "aum_trend_score",
            "turnover_ratio",
            "top10_concentration_pct",
            "downside_capture_pct",
            "expense_trend_delta",
        ],
    },
    "exit": {
        "formula": "Σ(component × weight) / Σ(available weights) × 10 — needs portfolio context",
        "scale": "0-100 (higher = stronger exit signal)",
        "weights": {
            "overlap": 25,
            "tax_impact": 25,
            "quality_inverse": 25,
            "cost": 15,
            "portfolio_fit": 10,
        },
        "component_logic": {
            "overlap": "overlap_pct/10, clamped [0,10]. 100% overlap → 10.",
            "tax_impact": "Inverse of tax / benefit ratio. Higher tax vs benefit "
                          "lowers exit desirability.",
            "quality_inverse": "10 − quality_0_10. Weak funds get higher exit scores.",
            "cost": "10 − er_regular × 5 (inverted sign for exit).",
            "portfolio_fit": "10 − portfolio_fit_0_10. Poor-fit funds favoured for exit.",
        },
        "primitives_used": [
            "overlap_pct (portfolio-context)",
            "tax_liability_rs", "tax_benefit_rs",
            "quality_score",
            "expense_ratio_regular",
            "portfolio_fit_score",
        ],
    },
    "add": {
        "formula": "Σ(component × weight) / Σ(available weights) × 10",
        "scale": "0-100",
        "weights": {
            "gap_fit": 30,
            "low_overlap": 25,
            "quality": 20,
            "need": 15,
            "cost": 10,
        },
        "component_logic": {
            "gap_fit": "How well this fund fills an asset-allocation gap in the portfolio.",
            "low_overlap": "10 − overlap_pct/10. Less overlap → higher add score.",
            "quality": "quality_score / 10.",
            "need": "Portfolio need for this category.",
            "cost": "Same as Quality.cost.",
        },
        "primitives_used": ["quality_score", "expense_ratio_direct", "portfolio gap context"],
    },
    "switch": {
        "formula": "(ΔQuality + ΔOverlap_reduction + Cost_saving − Tax_cost) / scale",
        "scale": "numeric; ≥ 2.0 → recommend switch",
        "description": "Only evaluated for Regular-plan holdings; compares against Direct sibling.",
        "primitives_used": [
            "quality_score_direct", "quality_score_regular",
            "overlap_pct", "expense_ratio_direct", "expense_ratio_regular",
            "holding_value_rs", "stcg_rate", "ltcg_rate", "holding_age_days",
        ],
    },
}


# ── Helpers ──────────────────────────────────────────────────────────────
def _derive_amc(scheme_name: Optional[str]) -> Optional[str]:
    if not scheme_name:
        return None
    known = [
        "Aditya Birla Sun Life", "Aditya Birla SL", "Canara Robeco", "Parag Parikh",
        "ICICI Prudential", "ICICI Pru", "Nippon India", "Mirae Asset",
        "Motilal Oswal", "Franklin India", "Franklin", "PGIM India", "PGIM",
        "Bandhan", "Edelweiss", "Quant", "HDFC", "SBI", "Axis", "Kotak",
        "DSP", "UTI", "Tata", "Navi", "JM", "IDFC", "Invesco", "Mahindra",
    ]
    for amc in known:
        if scheme_name.lower().startswith(amc.lower()):
            return amc
    return scheme_name.split(" ")[0]


def _derive_plan_type(scheme_name: Optional[str], slug: Optional[str]) -> str:
    name = (scheme_name or "").lower()
    slg = (slug or "").lower()
    if "regular" in name or "regular" in slg:
        return "regular"
    if "direct" in name or "direct" in slg:
        return "direct"
    return "direct"  # default


def _fund_row_to_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    """Transform a PG row into a dashboard-ready fund object with per-component
    Quality & Health breakdowns recomputed live (so the UI always reflects
    current weights & logic)."""
    # Raw primitives needed by v3_scoring
    def _f(v):
        """Cast Decimal/asyncpg types to float so v3_scoring math works."""
        return float(v) if v is not None else None
    prim = {
        "aum_cr": _f(row.get("aum_cr")),
        "fund_age_years": _f(row.get("fund_age_years")),
        "ret_1y": _f(row.get("ret_1y")),
        "ret_3y": _f(row.get("ret_3y")),
        "ret_5y": _f(row.get("ret_5y")),
        "cat_avg_1y": _f(row.get("category_avg_1y")),
        "cat_avg_3y": _f(row.get("category_avg_3y")),
        "cat_avg_5y": _f(row.get("category_avg_5y")),
        "sharpe": _f(row.get("sharpe")),
        "sortino": _f(row.get("sortino")),
        "consistency_score": _f(row.get("consistency_score")),
        "max_drawdown_pct": _f(row.get("max_drawdown_pct")),
        "expense_ratio_direct": _f(row.get("expense_ratio_direct")),
        "expense_ratio_regular": _f(row.get("expense_ratio_regular")),
        "expense_ratio": _f(row.get("expense_ratio_direct") or row.get("expense_ratio_regular")),
        "manager_tenure_years": _f(row.get("manager_tenure_years")),
        "aum_trend_score": _f(row.get("aum_trend_score")),
        "turnover_ratio": _f(row.get("turnover_ratio")),
        "top10_concentration_pct": _f(row.get("top10_concentration_pct")),
        "downside_capture_pct": _f(row.get("downside_capture_pct")),
        "expense_trend_delta": _f(row.get("expense_trend_delta")),
        # Debt primitives (Moneycontrol-sourced)
        "credit_quality_score": _f(row.get("credit_quality_score")),
        "duration_risk_score": _f(row.get("duration_risk_score")),
        "ytm": _f(row.get("ytm")),
        "modified_duration": _f(row.get("modified_duration")),
        "investment_style": row.get("investment_style"),
        "moneycontrol_imid": row.get("moneycontrol_imid"),
    }
    quality = v3_scoring.compute_quality_score({**prim,
                                                 "category": row.get("category"),
                                                 "sub_category": row.get("sub_category")})
    health = v3_scoring.compute_health_score({**prim,
                                              "category": row.get("category"),
                                              "sub_category": row.get("sub_category")})
    slug = row.get("groww_slug") or ""
    source = "tickertape" if slug.startswith("tt_") else "groww"
    scheme_name = row.get("instrument_name")
    return {
        "instrument_id": str(row["instrument_id"]) if row.get("instrument_id") else None,
        "scheme_name": scheme_name,
        "amc_name": _derive_amc(scheme_name),
        "category": row.get("category"),
        "sub_category": row.get("sub_category"),
        "plan_type": _derive_plan_type(scheme_name, slug),
        "source": source,
        "slug": slug,
        "manager_name": row.get("manager_name"),
        "risk_label": row.get("risk_label"),
        "last_scraped_at": row.get("last_scraped_at").isoformat() if row.get("last_scraped_at") else None,
        "scores": {
            "quality": _f(row.get("quality_score")),
            "health": _f(row.get("health_score")),
            "exit": _f(row.get("exit_score_baseline")),
            "add": _f(row.get("add_score_baseline")),
        },
        "primitives": prim,
        "quality_breakdown": quality,
        "health_breakdown": health,
    }


async def _fetch_master_funds(limit: Optional[int] = None) -> List[Dict[str, Any]]:
    pool = await pg_client.get_pool()
    # Only select columns that actually exist. Returns & ratio primitives
    # live in other tables (mutual_fund_ratios / nav-analytics joins); we
    # surface them when present via the v3 composites, otherwise NULL.
    sql = """
        SELECT im.instrument_id, im.instrument_name,
               mfm.category, mfm.sub_category, mfm.groww_slug,
               mfm.risk_label, mfm.last_scraped_at,
               mfm.quality_score, mfm.health_score,
               mfm.exit_score_baseline, mfm.add_score_baseline,
               mfm.aum_cr, mfm.fund_age_years,
               mfm.expense_ratio_direct, mfm.expense_ratio_regular,
               mfm.manager_tenure_years, mfm.manager_name,
               mfm.top10_concentration_pct, mfm.turnover_ratio,
               mfm.max_drawdown_pct, mfm.downside_capture_pct,
               mfm.consistency_score, mfm.aum_trend_score,
               mfm.expense_trend_delta,
               mfm.category_avg_1y, mfm.category_avg_3y, mfm.category_avg_5y,
               mfm.credit_quality_score, mfm.duration_risk_score,
               mfm.ytm, mfm.modified_duration,
               mfm.investment_style, mfm.moneycontrol_imid,
               mfr.ret_1y, mfr.ret_3y, mfr.ret_5y,
               mfr.sharpe, mfr.sortino
        FROM mutual_fund_metadata mfm
        JOIN instrument_master im ON im.instrument_id = mfm.instrument_id
        LEFT JOIN LATERAL (
            SELECT ret_1y, ret_3y, ret_5y, sharpe, sortino
            FROM mutual_fund_performance_ratios
            WHERE instrument_id = mfm.instrument_id
            ORDER BY ratios_date DESC LIMIT 1
        ) mfr ON TRUE
        WHERE im.is_active = TRUE
        ORDER BY mfm.quality_score DESC NULLS LAST,
                 im.instrument_name ASC
    """
    if limit:
        sql += f"\nLIMIT {int(limit)}"
    async with pool.acquire() as conn:
        rows = await conn.fetch(sql)
    return [_fund_row_to_payload(dict(r)) for r in rows]


# ── Endpoints ───────────────────────────────────────────────────────────
@router.get("/admin/v3-master-funds")
async def v3_master_funds(request: Request, limit: int = 0):
    """Return every scored mutual fund with primitives + Q/H sub-score breakdown."""
    await require_admin(request)
    try:
        await pg_client.get_pool()
    except Exception as e:  # noqa: BLE001
        logger.warning("v3-master-funds: pg unavailable → %s", e)
        raise HTTPException(status_code=503, detail="Postgres not configured") from e
    funds = await _fetch_master_funds(limit=limit if limit > 0 else None)
    by_cat: Dict[str, int] = {}
    by_source = {"groww": 0, "tickertape": 0}
    scored = 0
    for f in funds:
        cat = f.get("category") or "Unknown"
        by_cat[cat] = by_cat.get(cat, 0) + 1
        by_source[f["source"]] = by_source.get(f["source"], 0) + 1
        if f["scores"].get("quality") is not None:
            scored += 1
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "totals": {
            "total": len(funds),
            "scored": scored,
            "by_source": by_source,
            "by_category": by_cat,
        },
        "formula_reference": FORMULA_REFERENCE,
        "funds": funds,
    }


@router.get("/admin/v3-master-funds/export.xlsx")
async def v3_master_funds_export(request: Request):
    """Download a multi-sheet Excel workbook of all V3-scored master funds."""
    await require_admin(request)
    try:
        await pg_client.get_pool()
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=503, detail="Postgres not configured") from e

    funds = await _fetch_master_funds()

    # Excel generation
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}")

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CBD5E1")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = brd
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"

    def _autosize(ws, max_width=42):
        for col in ws.columns:
            length = 8
            letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    length = max(length, min(len(str(cell.value)) + 2, max_width))
            ws.column_dimensions[letter].width = length

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    summary_headers = [
        "Scheme name", "AMC", "Category", "Sub-category", "Plan", "Source",
        "Quality /100", "Health /100", "Exit /100", "Add /100",
        "AUM (₹ Cr)", "ER-Direct (%)", "ER-Regular (%)",
        "Mgr tenure (yrs)", "Fund age (yrs)",
        "Top-10 (%)", "Turnover (%)", "Max DD (%)",
        "Downside capture (%)", "Consistency",
        "Sharpe", "Sortino", "Last scraped",
    ]
    _write_header(ws, summary_headers)
    for i, f in enumerate(funds, 2):
        s, p = f["scores"], f["primitives"]
        row = [
            f["scheme_name"], f.get("amc_name"), f.get("category"), f.get("sub_category"),
            f.get("plan_type"), f["source"],
            s.get("quality"), s.get("health"), s.get("exit"), s.get("add"),
            p.get("aum_cr"), p.get("expense_ratio_direct"), p.get("expense_ratio_regular"),
            p.get("manager_tenure_years"), p.get("fund_age_years"),
            p.get("top10_concentration_pct"), p.get("turnover_ratio"),
            p.get("max_drawdown_pct"), p.get("downside_capture_pct"),
            p.get("consistency_score"),
            p.get("sharpe"), p.get("sortino"),
            f.get("last_scraped_at"),
        ]
        for col, v in enumerate(row, 1):
            ws.cell(row=i, column=col, value=v).border = brd
    _autosize(ws)

    # ── Sheet 2: All Primitives (wide dense grid) ──
    ws = wb.create_sheet("Primitives")
    all_prim_keys = [
        "aum_cr", "fund_age_years",
        "expense_ratio_direct", "expense_ratio_regular", "expense_trend_delta",
        "manager_tenure_years",
        "aum_trend_score", "turnover_ratio",
        "top10_concentration_pct",
        "downside_capture_pct", "consistency_score", "max_drawdown_pct",
        "sharpe", "sortino",
        "ret_1y", "ret_3y", "ret_5y",
        "cat_avg_1y", "cat_avg_3y", "cat_avg_5y",
    ]
    prim_headers = ["Scheme name", "Source", "Category", "Plan"] + all_prim_keys
    _write_header(ws, prim_headers)
    for i, f in enumerate(funds, 2):
        p = f["primitives"]
        row = [f["scheme_name"], f["source"], f.get("category"), f.get("plan_type")]
        row += [p.get(k) for k in all_prim_keys]
        for col, v in enumerate(row, 1):
            ws.cell(row=i, column=col, value=v).border = brd
    _autosize(ws)

    # ── Sheet 3: Quality breakdown ──
    ws = wb.create_sheet("Quality breakdown")
    q_comps = ["performance", "risk_adjusted", "consistency", "drawdown", "cost", "aum_age"]
    q_headers = ["Scheme name", "Source", "Quality /100"]
    for c in q_comps:
        q_headers += [f"{c} /10", f"{c} weight %"]
    q_headers += ["Missing primitives"]
    _write_header(ws, q_headers)
    for i, f in enumerate(funds, 2):
        qb = f["quality_breakdown"]
        comps = qb.get("components") or {}
        eff = qb.get("effective_weights") or {}
        row = [f["scheme_name"], f["source"], f["scores"].get("quality")]
        for c in q_comps:
            row += [comps.get(c), eff.get(c)]
        row += [", ".join(qb.get("missing_primitives") or [])]
        for col, v in enumerate(row, 1):
            ws.cell(row=i, column=col, value=v).border = brd
    _autosize(ws)

    # ── Sheet 4: Health breakdown ──
    ws = wb.create_sheet("Health breakdown")
    h_comps = ["manager_tenure", "aum_stability", "turnover", "concentration",
               "downside_protection", "expense_trend"]
    h_headers = ["Scheme name", "Source", "Health /100"]
    for c in h_comps:
        h_headers += [f"{c} /10", f"{c} weight %"]
    h_headers += ["Missing primitives"]
    _write_header(ws, h_headers)
    for i, f in enumerate(funds, 2):
        hb = f["health_breakdown"]
        comps = hb.get("components") or {}
        eff = hb.get("effective_weights") or {}
        row = [f["scheme_name"], f["source"], f["scores"].get("health")]
        for c in h_comps:
            row += [comps.get(c), eff.get(c)]
        row += [", ".join(hb.get("missing_primitives") or [])]
        for col, v in enumerate(row, 1):
            ws.cell(row=i, column=col, value=v).border = brd
    _autosize(ws)

    # ── Sheet 5: Formula reference ──
    ws = wb.create_sheet("Formula reference")
    r = 1
    ws.cell(row=r, column=1, value="V3 Engine — score formulas and primitive sources").font = Font(bold=True, size=13)
    r += 2
    for score_name, spec in FORMULA_REFERENCE.items():
        cell = ws.cell(row=r, column=1, value=score_name.upper())
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = hdr_fill
        r += 1
        ws.cell(row=r, column=1, value="Formula:").font = Font(bold=True)
        ws.cell(row=r, column=2, value=spec.get("formula"))
        r += 1
        ws.cell(row=r, column=1, value="Scale:").font = Font(bold=True)
        ws.cell(row=r, column=2, value=spec.get("scale"))
        r += 1
        if spec.get("weights"):
            ws.cell(row=r, column=1, value="Weights:").font = Font(bold=True)
            ws.cell(row=r, column=2, value=" · ".join(
                f"{k} {v}%" for k, v in spec["weights"].items()))
            r += 1
        if spec.get("component_logic"):
            ws.cell(row=r, column=1, value="Component logic:").font = Font(bold=True)
            r += 1
            for k, v in spec["component_logic"].items():
                c1 = ws.cell(row=r, column=1, value=f"  · {k}")
                c1.alignment = Alignment(indent=1)
                ws.cell(row=r, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
                r += 1
        if spec.get("primitives_used"):
            ws.cell(row=r, column=1, value="Primitives used:").font = Font(bold=True)
            ws.cell(row=r, column=2, value=", ".join(spec["primitives_used"])).alignment = Alignment(wrap_text=True)
            r += 1
        r += 1
    # Widen for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"v3-master-funds-{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
