"""
NIVESH QA — Test Case Generator
Produces: functional_test_cases.xlsx, functional_test_cases.csv,
          requirement_traceability_matrix.csv
Run: python qa/generate_test_cases.py
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Colour palette ────────────────────────────────────────────────────────────
COLOUR = {
    "critical":  "C62828",
    "high":      "E65100",
    "medium":    "F9A825",
    "low":       "2E7D32",
    "header_bg": "1A237E",
    "header_fg": "FFFFFF",
    "alt_row":   "EEF2FF",
}

HEADERS = [
    "TC ID", "Module", "Req ID", "Title", "Priority", "Severity",
    "Preconditions", "Test Data", "Steps (Given / When / Then)",
    "Expected Result", "Actual Result", "Status",
    "Automation Candidate", "Comments",
]

COL_WIDTHS = [14, 12, 16, 40, 10, 10, 35, 28, 60, 45, 15, 10, 14, 30]

# ── Test-case definitions ─────────────────────────────────────────────────────
# Each dict: tc_id, module, req_id, title, priority, severity,
#            preconditions, test_data, steps, expected, automation, comments
TEST_CASES = [

    # ══════════════════════════════════════════════════════════════════════════
    # AUTH
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_AUTH_001", mod="AUTH", req="FR-V1B-AUTH-001",
         title="Whitelisted user logs in via Google OAuth",
         pri="Critical", sev="Critical",
         pre="User email is in whitelisted_users collection; Google GIS SDK available",
         td="td_user_ahaan (ahaanporwal16@gmail.com)",
         steps=(
             "Given the user is on the landing page and Google GIS SDK is loaded\n"
             "When POST /api/auth/google is called with a valid Google id_token\n"
             "Then HTTP 200 is returned, an HttpOnly session_token cookie is set, "
             "and the response body contains user.email and user.role"
         ),
         exp="HTTP 200; session_token cookie present; user.email = ahaanporwal16@gmail.com; user.role = user",
         auto="Yes", cmt="Core auth — must pass pre-release"),

    dict(tc="TC_AUTH_002", mod="AUTH", req="FR-V1B-AUTH-002",
         title="Non-whitelisted user is rejected",
         pri="Critical", sev="Critical",
         pre="Email is NOT in whitelisted_users",
         td="random_user@gmail.com",
         steps=(
             "Given the user provides a valid Google id_token for a non-whitelisted email\n"
             "When POST /api/auth/google is called\n"
             "Then HTTP 403 is returned with message 'Access restricted'"
         ),
         exp="HTTP 403; body = {detail: 'Access restricted'}; no session cookie",
         auto="Yes", cmt="Whitelist enforcement"),

    dict(tc="TC_AUTH_003", mod="AUTH", req="FR-V1B-AUTH-001",
         title="Invalid or expired Google token rejected",
         pri="Critical", sev="Critical",
         pre="None",
         td="Malformed JWT string 'bad.token.here'",
         steps=(
             "Given no active session exists\n"
             "When POST /api/auth/google is called with an invalid id_token\n"
             "Then HTTP 401 is returned"
         ),
         exp="HTTP 401; no session cookie set",
         auto="Yes", cmt=""),

    dict(tc="TC_AUTH_004", mod="AUTH", req="FR-V1B-AUTH-003",
         title="GET /me returns current user for valid session",
         pri="Critical", sev="Critical",
         pre="User is logged in (valid session cookie)",
         td="td_user_ahaan",
         steps=(
             "Given the user has a valid session_token cookie\n"
             "When GET /api/auth/me is called\n"
             "Then HTTP 200 is returned with user object containing email, role, onboarding_completed"
         ),
         exp="HTTP 200; user.email matches; user.role = user",
         auto="Yes", cmt=""),

    dict(tc="TC_AUTH_005", mod="AUTH", req="FR-V1B-AUTH-003",
         title="GET /me returns 401 without session cookie",
         pri="Critical", sev="Critical",
         pre="No session cookie",
         td="No cookie",
         steps=(
             "Given no session_token cookie is present\n"
             "When GET /api/auth/me is called\n"
             "Then HTTP 401 is returned"
         ),
         exp="HTTP 401",
         auto="Yes", cmt=""),

    dict(tc="TC_AUTH_006", mod="AUTH", req="FR-V1B-AUTH-004",
         title="Logout clears session cookie",
         pri="High", sev="High",
         pre="User is logged in",
         td="td_user_ahaan",
         steps=(
             "Given the user has an active session\n"
             "When POST /api/auth/logout is called\n"
             "Then HTTP 200 is returned and the session_token cookie is cleared (Max-Age=0)"
         ),
         exp="HTTP 200; Set-Cookie header clears session_token",
         auto="Yes", cmt=""),

    dict(tc="TC_AUTH_007", mod="AUTH", req="FR-V1B-AUTH-005",
         title="Admin-only endpoint returns 403 for user role",
         pri="Critical", sev="Critical",
         pre="Logged in as role=user",
         td="td_user_ahaan",
         steps=(
             "Given the user is authenticated with role=user\n"
             "When GET /api/admin/users is called\n"
             "Then HTTP 403 is returned"
         ),
         exp="HTTP 403; body contains permission error",
         auto="Yes", cmt="RBAC enforcement"),

    dict(tc="TC_AUTH_008", mod="AUTH", req="FR-V1B-AUTH-006",
         title="Rate limiter blocks after 30 requests/minute",
         pri="High", sev="High",
         pre="None",
         td="Rapid-fire 31 POST /api/auth/google calls within 60s",
         steps=(
             "Given the API is running with rate limiting enabled\n"
             "When 31 POST /api/auth/google requests are sent within 60 seconds from the same IP\n"
             "Then the 31st request returns HTTP 429"
         ),
         exp="HTTP 429 on 31st request; Retry-After header present",
         auto="Yes", cmt="Sliding-window in-memory limiter"),

    # ══════════════════════════════════════════════════════════════════════════
    # ONBOARD
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_ONBOARD_001", mod="ONBOARD", req="FR-V1B-ONBOARD-001",
         title="New user selects journey type",
         pri="High", sev="High",
         pre="User authenticated, onboarding_completed=false",
         td="td_user_ahaan; journey_type=investor",
         steps=(
             "Given a newly authenticated user has not completed onboarding\n"
             "When PUT /api/user/journey with body {journey_type: 'investor'} is called\n"
             "Then HTTP 200 is returned and user.journey_type is persisted"
         ),
         exp="HTTP 200; user_profiles.journey_type = 'investor'",
         auto="Yes", cmt=""),

    dict(tc="TC_ONBOARD_002", mod="ONBOARD", req="FR-V1B-ONBOARD-002",
         title="Risk questionnaire computes correct category",
         pri="Critical", sev="High",
         pre="User authenticated",
         td="Answers: [3,3,3,3,3,3] — should yield 'moderate'",
         steps=(
             "Given the user is on the risk profile questionnaire\n"
             "When POST /api/user/risk-profile with 6 answers summing to 18 is called\n"
             "Then HTTP 200 is returned, risk_category = 'moderate', score = 18"
         ),
         exp="risk_category = moderate; score = 18",
         auto="Yes", cmt="6-question scoring formula"),

    dict(tc="TC_ONBOARD_003", mod="ONBOARD", req="FR-V1B-ONBOARD-002",
         title="Aggressive risk profile boundary (max score)",
         pri="High", sev="High",
         pre="User authenticated",
         td="Answers: [5,5,5,5,5,5] — score=30",
         steps=(
             "Given the user answers all 6 questions with score 5\n"
             "When POST /api/user/risk-profile is called\n"
             "Then risk_category = 'aggressive'"
         ),
         exp="risk_category = aggressive; score = 30",
         auto="Yes", cmt="Boundary: max aggressive"),

    dict(tc="TC_ONBOARD_004", mod="ONBOARD", req="FR-V1B-ONBOARD-003",
         title="Quick setup SIP projection for new investor",
         pri="High", sev="Medium",
         pre="User authenticated, journey_type=new_investor",
         td="monthly_sip=25000, horizon_years=20, risk=aggressive",
         steps=(
             "Given a new investor completes quick setup with SIP=₹25,000 and 20-year horizon\n"
             "When POST /api/user/quick-setup is called\n"
             "Then HTTP 200 is returned with projected_value > 0 computed via FV formula"
         ),
         exp="HTTP 200; projected_value > 0; allocation_equity ≥ 0.75 for aggressive",
         auto="Yes", cmt="FV formula validation"),

    dict(tc="TC_ONBOARD_005", mod="ONBOARD", req="FR-V1B-ONBOARD-004",
         title="Dashboard is gated until onboarding complete",
         pri="High", sev="High",
         pre="User authenticated, onboarding_completed=false",
         td="td_user_ahaan with onboarding_completed=false",
         steps=(
             "Given a user who has not completed onboarding\n"
             "When GET /api/portfolio is called\n"
             "Then HTTP 403 or redirect response with message about incomplete onboarding"
         ),
         exp="HTTP 403 or 302; message indicates onboarding required",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # PORTFOLIO
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_PORTFOLIO_001", mod="PORTFOLIO", req="FR-V1B-PORT-001",
         title="Create new portfolio",
         pri="Critical", sev="High",
         pre="User authenticated with onboarding_completed=true",
         td="td_user_ahaan; name='Family Portfolio'",
         steps=(
             "Given the user is authenticated and onboarding is complete\n"
             "When POST /api/portfolios with body {name: 'Family Portfolio'} is called\n"
             "Then HTTP 201 is returned with portfolio_id, name, and created_at"
         ),
         exp="HTTP 201; portfolio.name = 'Family Portfolio'; portfolio_id is UUID",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_002", mod="PORTFOLIO", req="FR-V1B-PORT-002",
         title="CAS PDF upload extracts holdings (CAMS provider)",
         pri="Critical", sev="Critical",
         pre="CAMS CAS PDF available; casparser.in reachable",
         td="td_cas_cams (8 MF holdings)",
         steps=(
             "Given the user has a valid CAMS CAS PDF\n"
             "When POST /api/portfolios/{id}/cas-upload with the PDF is called\n"
             "Then HTTP 200 is returned, holdings array has 8 items with isin, units, avg_nav"
         ),
         exp="HTTP 200; len(holdings) = 8; each holding has isin, units, avg_nav, folio_no",
         auto="Yes", cmt="3-provider fallback: Google DocAI → Claude Vision → casparser.in"),

    dict(tc="TC_PORTFOLIO_003", mod="PORTFOLIO", req="FR-V1B-PORT-002",
         title="CAS PDF upload extracts holdings (NSDL provider)",
         pri="Critical", sev="Critical",
         pre="NSDL CAS PDF available",
         td="td_cas_nsdl (15 holdings)",
         steps=(
             "Given the user has a valid NSDL CAS PDF\n"
             "When POST /api/portfolios/{id}/cas-upload with the NSDL PDF is called\n"
             "Then HTTP 200 is returned with 15 holdings"
         ),
         exp="HTTP 200; len(holdings) = 15; source = 'nsdl'",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_004", mod="PORTFOLIO", req="FR-V1B-PORT-002",
         title="Invalid PDF returns 422",
         pri="High", sev="High",
         pre="None",
         td="A plain text file renamed to .pdf",
         steps=(
             "Given the user uploads a non-CAS file\n"
             "When POST /api/portfolios/{id}/cas-upload is called with the invalid file\n"
             "Then HTTP 422 is returned with error detail"
         ),
         exp="HTTP 422; detail contains 'Could not extract holdings'",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_005", mod="PORTFOLIO", req="FR-V1B-PORT-003",
         title="Add holding manually",
         pri="High", sev="High",
         pre="Portfolio exists; instrument present in instrument_master",
         td="isin=INF082J01157 (HDFC Flexi Cap); units=100; avg_nav=850",
         steps=(
             "Given the user has an existing portfolio\n"
             "When POST /api/portfolios/{id}/holdings with isin, units, avg_nav, buy_date is called\n"
             "Then HTTP 201 is returned with holding_id"
         ),
         exp="HTTP 201; holding.isin = INF082J01157; holding.units = 100",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_006", mod="PORTFOLIO", req="FR-V1B-PORT-003",
         title="Edit holding updates buy_date and re-classifies tax",
         pri="High", sev="High",
         pre="Holding exists with buy_date 3 months ago",
         td="td_holding_stcg; new buy_date = 2 years ago",
         steps=(
             "Given a holding bought 3 months ago (STCG territory)\n"
             "When PATCH /api/holdings/{id} with buy_date = 2 years ago is called\n"
             "Then holding.holding_period changes and tax_classification becomes 'LTCG'"
         ),
         exp="HTTP 200; tax_classification = LTCG; holding_period > 365",
         auto="Yes", cmt="Buy-date reclassification"),

    dict(tc="TC_PORTFOLIO_007", mod="PORTFOLIO", req="FR-V1B-PORT-003",
         title="Delete holding removes from portfolio",
         pri="Medium", sev="Medium",
         pre="Holding exists in portfolio",
         td="td_user_ahaan; any holding_id",
         steps=(
             "Given a portfolio with at least one holding\n"
             "When DELETE /api/holdings/{id} is called\n"
             "Then HTTP 204 is returned and the holding no longer appears in GET /holdings"
         ),
         exp="HTTP 204; holding absent from subsequent GET",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_008", mod="PORTFOLIO", req="FR-V1B-PORT-004",
         title="Instrument search returns fuzzy matches",
         pri="Medium", sev="Medium",
         pre="instrument_master populated with ≥735 instruments",
         td="query='hdfc flexi'",
         steps=(
             "Given instrument_master has HDFC Flexi Cap Fund\n"
             "When GET /api/instruments/search?q=hdfc+flexi is called\n"
             "Then results include INF082J01157 in top 5"
         ),
         exp="HTTP 200; results[0].isin = INF082J01157 or in top 5",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_009", mod="PORTFOLIO", req="FR-V1B-PORT-005",
         title="CSV export contains all holdings",
         pri="Medium", sev="Medium",
         pre="Portfolio has ≥1 holding",
         td="td_user_ahaan portfolio with 5 holdings",
         steps=(
             "Given a portfolio with 5 holdings\n"
             "When GET /api/portfolios/{id}/export.csv is called\n"
             "Then a CSV file is returned with header row and 5 data rows"
         ),
         exp="Content-Type: text/csv; 6 rows (1 header + 5 data); columns include isin, name, units, value",
         auto="Yes", cmt=""),

    dict(tc="TC_PORTFOLIO_010", mod="PORTFOLIO", req="FR-V1B-PORT-006",
         title="Snapshot creation stores point-in-time portfolio",
         pri="High", sev="High",
         pre="Portfolio has holdings; snapshot_date in past",
         td="snapshot_date = 2025-03-31",
         steps=(
             "Given a portfolio with holdings\n"
             "When POST /api/portfolios/{id}/snapshots with snapshot_date=2025-03-31 is called\n"
             "Then a snapshot document is created with total_value at that date"
         ),
         exp="HTTP 201; snapshot.snapshot_date = 2025-03-31; snapshot.total_value > 0",
         auto="Yes", cmt="Time-machine feature"),

    # ══════════════════════════════════════════════════════════════════════════
    # INSIGHTS
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_INSIGHTS_001", mod="INSIGHTS", req="FR-V1B-INS-001",
         title="Deterministic insights return same result for same input",
         pri="Critical", sev="Critical",
         pre="Portfolio with holdings loaded",
         td="td_portfolio_overlap",
         steps=(
             "Given a portfolio with fixed holdings\n"
             "When GET /api/portfolios/{id}/insights is called twice with the same data\n"
             "Then both responses return identical insight list and values"
         ),
         exp="Response 1 == Response 2 (deterministic); insights array non-empty",
         auto="Yes", cmt="Determinism is a hard requirement"),

    dict(tc="TC_INSIGHTS_002", mod="INSIGHTS", req="FR-V1B-INS-001",
         title="Overlap insight fires for 70% fund overlap",
         pri="High", sev="High",
         pre="Portfolio has 2 large-cap funds with 70% overlap",
         td="td_portfolio_overlap",
         steps=(
             "Given a portfolio with 2 funds sharing 70% stock overlap\n"
             "When GET /api/portfolios/{id}/insights is called\n"
             "Then an OVERLAP insight is present with overlap_pct ≥ 70"
         ),
         exp="insight.type = OVERLAP; overlap_pct ≥ 70; severity = WARNING or CRITICAL",
         auto="Yes", cmt=""),

    dict(tc="TC_INSIGHTS_003", mod="INSIGHTS", req="FR-V1B-INS-001",
         title="AMC concentration insight fires at 60% HDFC exposure",
         pri="High", sev="High",
         pre="Portfolio has 60% allocation to HDFC AMC",
         td="td_portfolio_amc_conc",
         steps=(
             "Given a portfolio where 60% of AUM is in HDFC AMC funds\n"
             "When GET /api/portfolios/{id}/insights is called\n"
             "Then an AMC_CONCENTRATION insight is returned"
         ),
         exp="insight.type = AMC_CONCENTRATION; amc = HDFC; pct ≥ 60",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # GOALS
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_GOALS_001", mod="GOALS", req="FR-V1B-GOALS-001",
         title="Create retirement goal with Monte Carlo projection",
         pri="High", sev="High",
         pre="User authenticated, onboarding complete",
         td="td_goals_retirement (target=₹5Cr, horizon=25yr, sip=₹25K/mo)",
         steps=(
             "Given the user is on the goals page\n"
             "When POST /api/goals with target=50000000, horizon_years=25, monthly_sip=25000 is called\n"
             "Then HTTP 201 is returned with goal_id and probability_of_success (0–100)"
         ),
         exp="HTTP 201; goal.probability_of_success between 0 and 100; goal.type = retirement",
         auto="Yes", cmt="Monte Carlo projection"),

    dict(tc="TC_GOALS_002", mod="GOALS", req="FR-V1B-GOALS-002",
         title="Maximum 4 goals enforced",
         pri="High", sev="High",
         pre="User already has 4 goals",
         td="5th goal creation attempt",
         steps=(
             "Given a user already has 4 active goals\n"
             "When POST /api/goals is called for a 5th goal\n"
             "Then HTTP 422 is returned with message 'Maximum 4 goals allowed'"
         ),
         exp="HTTP 422; detail = 'Maximum 4 goals allowed'",
         auto="Yes", cmt="Max-4 limit is a hard constraint"),

    dict(tc="TC_GOALS_003", mod="GOALS", req="FR-V1B-GOALS-001",
         title="Goal probability updates when SIP amount changes",
         pri="Medium", sev="Medium",
         pre="Goal exists",
         td="Increase monthly_sip from ₹25K to ₹50K",
         steps=(
             "Given an existing goal with probability_of_success = 65%\n"
             "When PATCH /api/goals/{id} with monthly_sip=50000 is called\n"
             "Then probability_of_success increases"
         ),
         exp="New probability_of_success > 65",
         auto="Yes", cmt=""),

    dict(tc="TC_GOALS_004", mod="GOALS", req="FR-V1B-GOALS-003",
         title="Delete goal removes it from list",
         pri="Medium", sev="Medium",
         pre="Goal exists",
         td="Any goal_id",
         steps=(
             "Given an existing goal\n"
             "When DELETE /api/goals/{id} is called\n"
             "Then HTTP 204 is returned and goal absent in GET /goals"
         ),
         exp="HTTP 204; goal absent",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # BROKER
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_BROKER_001", mod="BROKER", req="FR-V1B-BROKER-001",
         title="Zerodha OAuth connect redirects to Kite",
         pri="High", sev="High",
         pre="ZERODHA_API_KEY configured in secrets",
         td="td_user_ahaan",
         steps=(
             "Given the user is authenticated and Zerodha API key is configured\n"
             "When GET /api/broker/zerodha/connect is called\n"
             "Then HTTP 302 redirect to Kite OAuth URL is returned"
         ),
         exp="HTTP 302; Location header points to kite.zerodha.com/connect/login",
         auto="Yes", cmt="Read-only broker integration"),

    dict(tc="TC_BROKER_002", mod="BROKER", req="FR-V1B-BROKER-001",
         title="Broker callback imports holdings into portfolio",
         pri="High", sev="High",
         pre="Valid broker callback token",
         td="Zerodha request_token from mock callback",
         steps=(
             "Given a valid Zerodha request_token from OAuth callback\n"
             "When GET /api/broker/zerodha/callback?request_token=<token> is called\n"
             "Then holdings are fetched and merged into the user's active portfolio"
         ),
         exp="HTTP 200; holdings added to portfolio; source = 'zerodha'",
         auto="No", cmt="Requires live Zerodha sandbox"),

    dict(tc="TC_BROKER_003", mod="BROKER", req="FR-V1B-BROKER-001",
         title="Broker disconnect clears token",
         pri="Medium", sev="Medium",
         pre="Broker connected",
         td="td_user_ahaan with active broker session",
         steps=(
             "Given the user has a connected Zerodha broker session\n"
             "When DELETE /api/broker/zerodha is called\n"
             "Then HTTP 200 is returned and broker token is removed"
         ),
         exp="HTTP 200; broker_token absent from user record",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # COMPLIANCE
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_COMPLIANCE_001", mod="COMPLIANCE", req="FR-V1B-COMP-001",
         title="Data export returns all user data (DPDP Article 5)",
         pri="High", sev="High",
         pre="User authenticated",
         td="td_user_ahaan",
         steps=(
             "Given an authenticated user\n"
             "When GET /api/compliance/data-export is called\n"
             "Then a JSON file with all user data (profile, holdings, goals, chat) is returned"
         ),
         exp="HTTP 200; Content-Type: application/json; includes portfolio, goals, chat_history",
         auto="Yes", cmt="DPDP compliance"),

    dict(tc="TC_COMPLIANCE_002", mod="COMPLIANCE", req="FR-V1B-COMP-002",
         title="Account deletion purges all PII",
         pri="High", sev="Critical",
         pre="User authenticated",
         td="td_user_ahaan",
         steps=(
             "Given an authenticated user\n"
             "When DELETE /api/compliance/account is called\n"
             "Then HTTP 200 is returned and all user documents are deleted from MongoDB"
         ),
         exp="HTTP 200; subsequent GET /me returns 401; MongoDB user docs absent",
         auto="Yes", cmt="Right-to-erasure"),

    dict(tc="TC_COMPLIANCE_003", mod="COMPLIANCE", req="FR-V1B-COMP-003",
         title="Consent record is stored on first login",
         pri="High", sev="Medium",
         pre="New user logging in for first time",
         td="td_user_ahaan first login",
         steps=(
             "Given a new user authenticates for the first time\n"
             "When POST /api/auth/google is called\n"
             "Then a consent record is created with timestamp and version"
         ),
         exp="consent_records collection has entry for user; consent.version = current",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # V3SCORE
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_V3SCORE_001", mod="V3SCORE", req="FR-V2B-SCORE-001",
         title="V3 scoring determinism — same input yields same output",
         pri="Critical", sev="Critical",
         pre="NAV primitives loaded for test fund",
         td="isin=INF082J01157; fixed NAV history snapshot",
         steps=(
             "Given V3 scoring primitives are loaded for a specific fund\n"
             "When GET /api/v3/score/{isin} is called twice with identical NAV data\n"
             "Then both responses return identical quality_score, health_score, exit_score"
         ),
         exp="Call1.quality_score == Call2.quality_score (all 5 composites identical)",
         auto="Yes", cmt="Determinism is a hard requirement"),

    dict(tc="TC_V3SCORE_002", mod="V3SCORE", req="FR-V2B-SCORE-001",
         title="Quality score range 0–100",
         pri="Critical", sev="High",
         pre="V3 primitives available",
         td="10 different ISINs",
         steps=(
             "Given V3 scoring engine is running\n"
             "When GET /api/v3/score/{isin} is called for 10 different funds\n"
             "Then all quality_score values are between 0 and 100 inclusive"
         ),
         exp="0 ≤ quality_score ≤ 100 for all 10 funds",
         auto="Yes", cmt=""),

    dict(tc="TC_V3SCORE_003", mod="V3SCORE", req="FR-V2B-SCORE-002",
         title="Danger classification: Q<40 = CRITICAL",
         pri="Critical", sev="Critical",
         pre="Fund with quality_score < 40",
         td="isin with mocked quality_score=35",
         steps=(
             "Given a fund with quality_score = 35 (< 40 threshold)\n"
             "When GET /api/v3/score/{isin} is called\n"
             "Then danger_level = CRITICAL"
         ),
         exp="danger_level = CRITICAL",
         auto="Yes", cmt="Q<40 threshold"),

    dict(tc="TC_V3SCORE_004", mod="V3SCORE", req="FR-V2B-SCORE-002",
         title="Danger classification: H<40 = CRITICAL",
         pri="Critical", sev="Critical",
         pre="Fund with health_score < 40",
         td="isin with mocked health_score=30",
         steps=(
             "Given a fund with health_score = 30 (< 40)\n"
             "When GET /api/v3/score/{isin} is called\n"
             "Then danger_level = CRITICAL"
         ),
         exp="danger_level = CRITICAL",
         auto="Yes", cmt="H<40 threshold"),

    dict(tc="TC_V3SCORE_005", mod="V3SCORE", req="FR-V2B-SCORE-003",
         title="Switch score formula correctness",
         pri="Critical", sev="Critical",
         pre="Source and target fund scores available",
         td="Q_new=72, Q_old=45, overlap_reduction=30, cost_saving=15000, tax_cost=5000",
         steps=(
             "Given source fund quality=45 and target quality=72\n"
             "When POST /api/v3/switch-score with the fund pair is called\n"
             "Then switch_score = (72-45) + (30/10) + (15000/10000) - (5000/10000) = 27+3+1.5-0.5 = 31"
         ),
         exp="switch_score ≥ 2.0; raw = 31 (± 0.01 float tolerance)",
         auto="Yes", cmt="Formula: (Q_new-Q_old) + overlap_red + (cost/10K) - (tax/10K)"),

    dict(tc="TC_V3SCORE_006", mod="V3SCORE", req="FR-V2B-SCORE-004",
         title="NAV analytics sweep runs at 22:30 IST",
         pri="High", sev="High",
         pre="APScheduler configured; NAV data available",
         td="Scheduled job trigger",
         steps=(
             "Given APScheduler is running\n"
             "When the clock reaches 22:30 IST on a weekday\n"
             "Then the NAV analytics sweep job executes and last_sweep_at is updated"
         ),
         exp="job_status = completed; last_sweep_at within 60s of 22:30 IST",
         auto="No", cmt="Scheduler integration — manual verification"),

    dict(tc="TC_V3SCORE_007", mod="V3SCORE", req="FR-V2B-SCORE-001",
         title="Portfolio-fit score decreases with high AMC concentration",
         pri="High", sev="High",
         pre="Portfolio with 60% HDFC AMC",
         td="td_portfolio_amc_conc",
         steps=(
             "Given a portfolio where HDFC AMC = 60% of AUM\n"
             "When GET /api/v3/portfolio-score/{portfolio_id} is called\n"
             "Then portfolio_fit_score is reduced by concentration penalty"
         ),
         exp="portfolio_fit_score < 60 (penalty applied)",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # PLAN
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_PLAN_001", mod="PLAN", req="FR-V2B-PLAN-001",
         title="Action plan generation returns engine version v2.5",
         pri="Critical", sev="Critical",
         pre="Portfolio with holdings; V3 scores available",
         td="td_user_ahaan with td_portfolio_overlap",
         steps=(
             "Given a portfolio with holdings and V3 scores computed\n"
             "When POST /api/plans/generate with portfolio_id is called\n"
             "Then HTTP 200 is returned with plan.engine_version = 'v2.5'"
         ),
         exp="plan.engine_version = 'v2.5'; plan.status = 'preview'",
         auto="Yes", cmt="Version constant must be 'v2.5'"),

    dict(tc="TC_PLAN_002", mod="PLAN", req="FR-V2B-PLAN-002",
         title="Rule 1: Regular→Direct duplicate detected",
         pri="Critical", sev="Critical",
         pre="Portfolio has HDFC Flexi Cap Regular AND Direct",
         td="td_portfolio_regular_direct",
         steps=(
             "Given a portfolio containing HDFC Flexi Cap Regular and HDFC Flexi Cap Direct\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains a REGULAR_TO_DIRECT action for the regular fund"
         ),
         exp="action.rule = RULE_1_REGULAR_DIRECT; action.fund = HDFC Flexi Cap Regular",
         auto="Yes", cmt="Fund name normalisation required"),

    dict(tc="TC_PLAN_003", mod="PLAN", req="FR-V2B-PLAN-003",
         title="Rule 2: AMC concentration action fires at >15%",
         pri="Critical", sev="High",
         pre="Single AMC exceeds 15% of portfolio",
         td="td_portfolio_amc_conc (HDFC = 60%)",
         steps=(
             "Given a portfolio with HDFC AMC at 60% of AUM\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains AMC_CONCENTRATION action targeting HDFC"
         ),
         exp="action.rule = RULE_2_AMC_CONC; action.amc = HDFC; action.current_pct = 60",
         auto="Yes", cmt=""),

    dict(tc="TC_PLAN_004", mod="PLAN", req="FR-V2B-PLAN-003b",
         title="Rule 2b: Category concentration fires at >35%",
         pri="High", sev="High",
         pre="Single SEBI category exceeds 35% of portfolio",
         td="Large Cap = 50% of portfolio",
         steps=(
             "Given a portfolio with 50% in Large Cap category\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains CATEGORY_CONCENTRATION action"
         ),
         exp="action.rule = RULE_2B_CAT_CONC; category = Large Cap; pct = 50",
         auto="Yes", cmt=""),

    dict(tc="TC_PLAN_005", mod="PLAN", req="FR-V2B-PLAN-004",
         title="Rule 3: Underperformer replacement (quality<6.5, ret_1y<8%)",
         pri="Critical", sev="High",
         pre="Fund with quality_score<6.5 AND 1yr return<8% in portfolio",
         td="Fund with quality=5.0, ret_1y=4%",
         steps=(
             "Given a portfolio containing a fund with quality=5.0 and 1yr return=4%\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains UNDERPERFORMER action with peer recommendation"
         ),
         exp="action.rule = RULE_3_UNDERPERFORMER; replacement_fund present",
         auto="Yes", cmt="quality<6.5 AND ret_1y<8% both required"),

    dict(tc="TC_PLAN_006", mod="PLAN", req="FR-V2B-PLAN-005",
         title="Rule 4: Overlap consolidation at >60% pairwise",
         pri="Critical", sev="High",
         pre="Portfolio has 2 funds with >60% overlap",
         td="td_portfolio_overlap (70% overlap)",
         steps=(
             "Given two funds in the portfolio share 70% stock overlap\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains OVERLAP action recommending consolidation"
         ),
         exp="action.rule = RULE_4_OVERLAP; overlap_pct ≥ 60",
         auto="Yes", cmt=""),

    dict(tc="TC_PLAN_007", mod="PLAN", req="FR-V2B-PLAN-006",
         title="Rule 5: Debt gap fires when debt < risk-profile floor",
         pri="Critical", sev="High",
         pre="User risk=aggressive, portfolio debt=0%",
         td="td_portfolio_debt_gap",
         steps=(
             "Given a user with risk_category=aggressive and 0% debt allocation\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains DEBT_GAP action with recommended debt allocation"
         ),
         exp="action.rule = RULE_5_DEBT_GAP; recommended_debt_pct > 0",
         auto="Yes", cmt=""),

    dict(tc="TC_PLAN_008", mod="PLAN", req="FR-V2B-PLAN-007",
         title="Rule 6: Cost-leak switch when annual_leak > ₹10,000",
         pri="High", sev="High",
         pre="Regular fund with annual cost leak > ₹10K",
         td="Regular fund with ₹15K annual leak",
         steps=(
             "Given a regular fund where annual expense leak = ₹15,000\n"
             "When POST /api/plans/generate is called\n"
             "Then plan contains COST_LEAK_SWITCH action"
         ),
         exp="action.rule = RULE_6_COST_LEAK; annual_leak ≥ 10000",
         auto="Yes", cmt="₹10K threshold"),

    dict(tc="TC_PLAN_009", mod="PLAN", req="FR-V2B-PLAN-008",
         title="Guardrail: High-Quality Protection blocks exit (Q≥75, H≥70)",
         pri="Critical", sev="Critical",
         pre="Fund with quality≥75 and health≥70 in plan exit action",
         td="Fund with Q=80, H=75",
         steps=(
             "Given an action plan exit action targets a fund with Q=80, H=75\n"
             "When plan generation runs guardrail check\n"
             "Then the exit action is blocked with reason HIGH_QUALITY_PROTECTION"
         ),
         exp="action.blocked = true; block_reason = HIGH_QUALITY_PROTECTION",
         auto="Yes", cmt="Guardrail 1"),

    dict(tc="TC_PLAN_010", mod="PLAN", req="FR-V2B-PLAN-008",
         title="Guardrail: Tax-Exceeds-Benefit blocks switch",
         pri="Critical", sev="Critical",
         pre="Switch where tax_cost > benefit",
         td="tax_cost=₹50K, projected_benefit=₹20K",
         steps=(
             "Given a switch action where tax cost=₹50K exceeds projected benefit=₹20K\n"
             "When guardrail check runs\n"
             "Then the switch is blocked with reason TAX_EXCEEDS_BENEFIT"
         ),
         exp="action.blocked = true; block_reason = TAX_EXCEEDS_BENEFIT",
         auto="Yes", cmt="Guardrail 2"),

    dict(tc="TC_PLAN_011", mod="PLAN", req="FR-V2B-PLAN-008",
         title="Guardrail: Recent-Investment Lockout (<6 months)",
         pri="Critical", sev="High",
         pre="Fund bought < 6 months ago",
         td="buy_date = 3 months ago",
         steps=(
             "Given a fund bought 3 months ago is targeted for an exit action\n"
             "When guardrail check runs\n"
             "Then the action is blocked with reason RECENT_INVESTMENT_LOCKOUT"
         ),
         exp="action.blocked = true; block_reason = RECENT_INVESTMENT_LOCKOUT",
         auto="Yes", cmt="Guardrail 3"),

    dict(tc="TC_PLAN_012", mod="PLAN", req="FR-V2B-PLAN-008",
         title="Guardrail: Low-Confidence blocks action (<50%)",
         pri="High", sev="High",
         pre="Action with confidence < 50%",
         td="action.confidence = 40%",
         steps=(
             "Given an action with confidence_score = 40%\n"
             "When guardrail check runs\n"
             "Then the action is blocked with reason LOW_CONFIDENCE"
         ),
         exp="action.blocked = true; block_reason = LOW_CONFIDENCE",
         auto="Yes", cmt="Guardrail 4"),

    dict(tc="TC_PLAN_013", mod="PLAN", req="FR-V2B-PLAN-009",
         title="Plan state machine: preview → active",
         pri="Critical", sev="High",
         pre="Plan in preview state",
         td="plan_id in preview",
         steps=(
             "Given a plan in 'preview' state\n"
             "When PATCH /api/plans/{id}/activate is called\n"
             "Then plan.status changes to 'active' and activated_at is set"
         ),
         exp="plan.status = active; plan.activated_at present",
         auto="Yes", cmt="State machine"),

    dict(tc="TC_PLAN_014", mod="PLAN", req="FR-V2B-PLAN-009",
         title="Plan state machine: active → archived",
         pri="High", sev="Medium",
         pre="Plan in active state",
         td="plan_id in active",
         steps=(
             "Given a plan in 'active' state\n"
             "When PATCH /api/plans/{id}/archive is called\n"
             "Then plan.status changes to 'archived'"
         ),
         exp="plan.status = archived",
         auto="Yes", cmt=""),

    dict(tc="TC_PLAN_015", mod="PLAN", req="FR-V2B-PLAN-009",
         title="Cannot activate a plan from archived state",
         pri="High", sev="Medium",
         pre="Plan in archived state",
         td="plan_id in archived",
         steps=(
             "Given a plan in 'archived' state\n"
             "When PATCH /api/plans/{id}/activate is called\n"
             "Then HTTP 422 is returned with invalid state transition message"
         ),
         exp="HTTP 422; detail = 'Cannot activate archived plan'",
         auto="Yes", cmt="Invalid state transition"),

    # ══════════════════════════════════════════════════════════════════════════
    # INTEL
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_INTEL_001", mod="INTEL", req="FR-V2B-INTEL-001",
         title="Look-through returns underlying stock exposure",
         pri="High", sev="High",
         pre="Fund with known holdings; fund_holdings data available",
         td="isin=INF082J01157 (HDFC Flexi Cap)",
         steps=(
             "Given a fund's underlying stock holdings are available\n"
             "When GET /api/intel/look-through?isin={isin} is called\n"
             "Then a list of underlying stocks with exposure percentages is returned"
         ),
         exp="HTTP 200; stocks array non-empty; sum of pct ≈ 100",
         auto="Yes", cmt="Portfolio look-through"),

    dict(tc="TC_INTEL_002", mod="INTEL", req="FR-V2B-INTEL-001",
         title="Portfolio-level sector exposure aggregation",
         pri="High", sev="Medium",
         pre="Portfolio with multiple funds",
         td="td_portfolio_overlap",
         steps=(
             "Given a portfolio with 3 funds\n"
             "When GET /api/intel/sector-exposure/{portfolio_id} is called\n"
             "Then aggregated sector allocations across all funds are returned"
         ),
         exp="HTTP 200; sectors array with pct; sum ≈ 100",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # HEALTH
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_HEALTH_001", mod="HEALTH", req="FR-V2B-HEALTH-001",
         title="Portfolio health score between 0 and 100",
         pri="High", sev="High",
         pre="Portfolio with holdings",
         td="td_user_ahaan",
         steps=(
             "Given a portfolio with holdings\n"
             "When GET /api/portfolios/{id}/health is called\n"
             "Then health_score is between 0 and 100"
         ),
         exp="0 ≤ health_score ≤ 100; component breakdown present",
         auto="Yes", cmt=""),

    dict(tc="TC_HEALTH_002", mod="HEALTH", req="FR-V2B-HEALTH-001",
         title="Empty portfolio returns health_score = 0",
         pri="Medium", sev="Low",
         pre="Portfolio exists but has no holdings",
         td="Empty portfolio",
         steps=(
             "Given a portfolio with no holdings\n"
             "When GET /api/portfolios/{id}/health is called\n"
             "Then health_score = 0 or null"
         ),
         exp="health_score = 0 or null; no error",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # TAX
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_TAX_001", mod="TAX", req="FR-V2B-TAX-001",
         title="Equity STCG computed at 20% (held < 12 months)",
         pri="Critical", sev="Critical",
         pre="Equity holding bought < 12 months ago",
         td="td_holding_stcg (bought 3 months ago, gain=₹10,000)",
         steps=(
             "Given an equity fund bought 3 months ago with ₹10,000 unrealized gain\n"
             "When GET /api/portfolios/{id}/tax-summary is called\n"
             "Then tax for this holding = ₹10,000 × 20% = ₹2,000"
         ),
         exp="holding.tax_type = STCG; holding.tax_rate = 0.20; holding.tax_amount = 2000",
         auto="Yes", cmt="FY2025-26 rate: Equity STCG = 20%"),

    dict(tc="TC_TAX_002", mod="TAX", req="FR-V2B-TAX-001",
         title="Equity LTCG computed at 12.5% above ₹1.25L exemption",
         pri="Critical", sev="Critical",
         pre="Equity holding bought > 12 months ago, gain > ₹1.25L",
         td="td_holding_ltcg (18 months, gain=₹2,00,000)",
         steps=(
             "Given an equity fund bought 18 months ago with ₹2,00,000 gain\n"
             "When GET /api/portfolios/{id}/tax-summary is called\n"
             "Then taxable gain = ₹75,000 (₹2L - ₹1.25L exemption); tax = ₹9,375"
         ),
         exp="holding.tax_type = LTCG; taxable_gain = 75000; tax_amount = 9375",
         auto="Yes", cmt="FY2025-26: LTCG 12.5% above ₹1.25L"),

    dict(tc="TC_TAX_003", mod="TAX", req="FR-V2B-TAX-001",
         title="Debt fund gains taxed at slab rate",
         pri="Critical", sev="Critical",
         pre="Debt fund holding with gain",
         td="Debt fund, gain=₹50,000, user slab=30%",
         steps=(
             "Given a debt fund with ₹50,000 gain\n"
             "When GET /api/portfolios/{id}/tax-summary is called\n"
             "Then tax_type = SLAB_RATE; tax_amount = ₹15,000"
         ),
         exp="tax_type = SLAB_RATE; tax_amount = 15000 (at 30% slab)",
         auto="Yes", cmt="Debt = slab rate regardless of holding period"),

    dict(tc="TC_TAX_004", mod="TAX", req="FR-V2B-TAX-001",
         title="ELSS lock-in enforced (3-year period)",
         pri="High", sev="High",
         pre="ELSS fund bought 2 years ago",
         td="td_holding_elss",
         steps=(
             "Given an ELSS fund bought 2 years ago (still in lock-in)\n"
             "When GET /api/portfolios/{id}/tax-summary is called\n"
             "Then the holding is flagged as locked_in = true"
         ),
         exp="holding.locked_in = true; lock_in_end_date = buy_date + 3 years",
         auto="Yes", cmt="ELSS 3-year lock-in"),

    dict(tc="TC_TAX_005", mod="TAX", req="FR-V2B-TAX-002",
         title="Switch cost calculator: full cost/benefit breakdown",
         pri="High", sev="High",
         pre="Source and target fund available",
         td="Source: regular fund; Target: direct equivalent",
         steps=(
             "Given a source regular fund and target direct fund\n"
             "When POST /api/tax/switch-cost with fund pair is called\n"
             "Then response includes exit_load, tax_cost, annual_saving, breakeven_months"
         ),
         exp="HTTP 200; exit_load ≥ 0; tax_cost ≥ 0; breakeven_months > 0",
         auto="Yes", cmt=""),

    dict(tc="TC_TAX_006", mod="TAX", req="FR-V2B-TAX-003",
         title="Tax harvesting: LTCG utilises ₹1.25L exemption",
         pri="High", sev="High",
         pre="Holdings with unrealized LTCG < ₹1.25L",
         td="Holdings with LTCG=₹80,000 (within exemption)",
         steps=(
             "Given holdings with unrealized LTCG = ₹80,000\n"
             "When GET /api/tax/harvesting-suggestions is called\n"
             "Then suggestion shows available exemption = ₹45,000 remaining"
         ),
         exp="available_exemption = 45000; suggestion.type = HARVEST_LTCG",
         auto="Yes", cmt="₹1.25L annual LTCG exemption"),

    dict(tc="TC_TAX_007", mod="TAX", req="FR-V2B-TAX-001",
         title="FIFO cost basis applied for partial redemption",
         pri="High", sev="High",
         pre="Multiple buy lots for same fund",
         td="Fund bought in 3 lots; partial redemption of 50 units",
         steps=(
             "Given a fund with 3 buy lots: 30 units @ ₹100, 20 units @ ₹110, 50 units @ ₹120\n"
             "When 50 units are redeemed\n"
             "Then cost basis uses FIFO: first 30 units @ ₹100, then 20 units @ ₹110"
         ),
         exp="cost_basis computed via FIFO; tax_computation uses oldest lots first",
         auto="Yes", cmt="FIFO is mandated by Indian tax law"),

    # ══════════════════════════════════════════════════════════════════════════
    # ENRICH
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_ENRICH_001", mod="ENRICH", req="FR-V2B-ENRICH-001",
         title="Enriched holdings API returns V3 scores + metadata",
         pri="Critical", sev="High",
         pre="Holdings exist; V3 scores computed",
         td="td_user_ahaan with holdings",
         steps=(
             "Given a portfolio with holdings and V3 scores computed\n"
             "When GET /api/portfolios/{id}/holdings/enriched is called\n"
             "Then each holding has quality_score, health_score, exit_score, danger_level"
         ),
         exp="HTTP 200; each holding has V3 scores; fund_name and amc populated",
         auto="Yes", cmt=""),

    dict(tc="TC_ENRICH_002", mod="ENRICH", req="FR-V2B-ENRICH-001",
         title="Enriched holdings returns current_value based on latest NAV",
         pri="High", sev="High",
         pre="Holdings with units; latest NAV in DB",
         td="100 units; NAV = ₹850",
         steps=(
             "Given 100 units of a fund with latest NAV = ₹850\n"
             "When GET /api/portfolios/{id}/holdings/enriched is called\n"
             "Then current_value = ₹85,000"
         ),
         exp="current_value = 85000; gain_loss = current_value - (units × avg_nav)",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # COPILOT
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_COPILOT_001", mod="COPILOT", req="FR-COP1-001",
         title="New chat session created",
         pri="Critical", sev="High",
         pre="User authenticated",
         td="td_user_ahaan",
         steps=(
             "Given an authenticated user\n"
             "When POST /api/copilot/sessions is called\n"
             "Then HTTP 201 is returned with session_id"
         ),
         exp="HTTP 201; session_id is UUID; session.user_id matches",
         auto="Yes", cmt=""),

    dict(tc="TC_COPILOT_002", mod="COPILOT", req="FR-COP1-002",
         title="Context warmup populates Redis cache",
         pri="High", sev="High",
         pre="User has portfolio, goals, plan",
         td="td_user_ahaan",
         steps=(
             "Given a user with portfolio, goals, and plan\n"
             "When POST /api/copilot/sessions/{id}/warmup is called\n"
             "Then context is cached in Redis with 5-minute TTL"
         ),
         exp="HTTP 200; Redis key exists; TTL ≤ 300s",
         auto="Yes", cmt="Sub-second first response requirement"),

    dict(tc="TC_COPILOT_003", mod="COPILOT", req="FR-COP1-003",
         title="Intent classification: portfolio question → PORTFOLIO intent",
         pri="Critical", sev="High",
         pre="Chat session exists",
         td="Message: 'What is my portfolio value?'",
         steps=(
             "Given a chat session\n"
             "When POST /api/copilot/sessions/{id}/classify with 'What is my portfolio value?' is called\n"
             "Then intent = PORTFOLIO; latency < 1ms"
         ),
         exp="intent = PORTFOLIO; classification_latency_ms < 1",
         auto="Yes", cmt="Deterministic keyword router, NOT LLM"),

    dict(tc="TC_COPILOT_004", mod="COPILOT", req="FR-COP1-003",
         title="Intent classification: tax question → TAX intent",
         pri="High", sev="High",
         pre="Chat session exists",
         td="Message: 'How much LTCG tax will I pay?'",
         steps=(
             "Given a chat session\n"
             "When the user sends 'How much LTCG tax will I pay?'\n"
             "Then intent = TAX"
         ),
         exp="intent = TAX",
         auto="Yes", cmt=""),

    dict(tc="TC_COPILOT_005", mod="COPILOT", req="FR-COP1-003",
         title="Intent classification: goal question → GOALS intent",
         pri="High", sev="Medium",
         pre="Chat session exists",
         td="Message: 'Am I on track for retirement?'",
         steps=(
             "Given a chat session\n"
             "When the user sends 'Am I on track for retirement?'\n"
             "Then intent = GOALS"
         ),
         exp="intent = GOALS",
         auto="Yes", cmt=""),

    dict(tc="TC_COPILOT_006", mod="COPILOT", req="FR-COP1-004",
         title="Context retrieval injects 5 blocks into LLM prompt",
         pri="Critical", sev="High",
         pre="User has portfolio, plan, goals, health, actions",
         td="td_user_ahaan",
         steps=(
             "Given a user with all context types populated\n"
             "When POST /api/copilot/sessions/{id}/chat is called with PORTFOLIO intent\n"
             "Then the LLM prompt contains all 5 context blocks: portfolio, plan, goals, health, actions"
         ),
         exp="LLM prompt has 5 context sections; no raw DB objects in prompt",
         auto="Yes", cmt="5 context blocks required"),

    dict(tc="TC_COPILOT_007", mod="COPILOT", req="FR-COP1-007",
         title="PAN number is redacted before reaching LLM",
         pri="Critical", sev="Critical",
         pre="PII scrubber active",
         td="td_chat_pan ('My PAN is ABCDE1234F')",
         steps=(
             "Given a user message containing PAN 'ABCDE1234F'\n"
             "When the message is processed by llm_safety.py\n"
             "Then the PAN is replaced with [REDACTED] before the OpenAI API call"
         ),
         exp="OpenAI request body does NOT contain 'ABCDE1234F'; contains '[REDACTED]'",
         auto="Yes", cmt="Critical PII safety — PAN/Aadhaar must never reach OpenAI"),

    dict(tc="TC_COPILOT_008", mod="COPILOT", req="FR-COP1-007",
         title="Aadhaar number is redacted before reaching LLM",
         pri="Critical", sev="Critical",
         pre="PII scrubber active",
         td="Message with Aadhaar: '1234 5678 9012'",
         steps=(
             "Given a user message containing Aadhaar number '1234 5678 9012'\n"
             "When processed by llm_safety.py\n"
             "Then the Aadhaar is replaced with [REDACTED]"
         ),
         exp="OpenAI request body contains '[REDACTED]' not '1234 5678 9012'",
         auto="Yes", cmt="Critical PII safety"),

    dict(tc="TC_COPILOT_009", mod="COPILOT", req="FR-COP1-006",
         title="Streaming SSE response emits tokens",
         pri="High", sev="High",
         pre="Chat session with warm context",
         td="Message: 'How is my portfolio doing?'",
         steps=(
             "Given a warm chat session\n"
             "When POST /api/copilot/sessions/{id}/chat?stream=true is called\n"
             "Then the response is text/event-stream with sequential token events"
         ),
         exp="Content-Type: text/event-stream; events have data: {token} format",
         auto="Yes", cmt="SSE streaming"),

    dict(tc="TC_COPILOT_010", mod="COPILOT", req="FR-COP1-006",
         title="Rate limit: 30 chat requests/minute enforced",
         pri="High", sev="High",
         pre="Active chat session",
         td="31 chat requests within 60 seconds",
         steps=(
             "Given an active chat session\n"
             "When 31 POST /api/copilot/sessions/{id}/chat calls are made within 60s\n"
             "Then the 31st call returns HTTP 429"
         ),
         exp="HTTP 429 on 31st call",
         auto="Yes", cmt=""),

    dict(tc="TC_COPILOT_011", mod="COPILOT", req="FR-COP1-005",
         title="Chart spec generated server-side for portfolio breakdown",
         pri="High", sev="High",
         pre="Portfolio with holdings",
         td="td_user_ahaan",
         steps=(
             "Given a portfolio\n"
             "When POST /api/copilot/charts/portfolio-breakdown is called\n"
             "Then chart spec JSON is returned (never LLM-generated)"
         ),
         exp="HTTP 200; spec.type in ['donut','bar','line','heatmap']; spec.data non-empty",
         auto="Yes", cmt="Chart spec is deterministic — never from LLM"),

    dict(tc="TC_COPILOT_012", mod="COPILOT", req="FR-COP1-008",
         title="Goal-level copilot responds to goal-specific questions",
         pri="High", sev="Medium",
         pre="Goal exists with projections",
         td="td_goals_retirement",
         steps=(
             "Given a retirement goal\n"
             "When POST /api/copilot/goals/{goal_id}/chat with 'Increase my SIP?' is called\n"
             "Then response is contextualised to the specific goal"
         ),
         exp="HTTP 200; response references goal.name or goal.target_amount",
         auto="No", cmt="Goal-level contextualisation"),

    dict(tc="TC_COPILOT_013", mod="COPILOT", req="FR-COP1-009",
         title="Scenario engine returns pre-built scenario cards",
         pri="Medium", sev="Medium",
         pre="User authenticated",
         td="td_user_ahaan",
         steps=(
             "Given an authenticated user\n"
             "When GET /api/copilot/scenarios is called\n"
             "Then a list of scenario cards is returned"
         ),
         exp="HTTP 200; scenarios array with title, description, prompt_template",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # COPILOT2 (V2 LangGraph)
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_COPILOT2_001", mod="COPILOT2", req="FR-COP2-001",
         title="LangGraph agent returns structured action response",
         pri="High", sev="High",
         pre="LangGraph dependencies installed; OpenAI key configured",
         td="td_user_ahaan; query='What should I do with my portfolio?'",
         steps=(
             "Given the Copilot V2 LangGraph agent is initialized\n"
             "When POST /api/copilot/v2/sessions/{id}/chat is called\n"
             "Then a structured response with action_plan recommendations is returned"
         ),
         exp="HTTP 200; response.actions non-empty; response.reasoning present",
         auto="No", cmt="LangGraph integration — requires OpenAI"),

    dict(tc="TC_COPILOT2_002", mod="COPILOT2", req="FR-COP2-002",
         title="SIP calculator tool returns monthly SIP for target amount",
         pri="High", sev="High",
         pre="SIP tool available in copilot_tools",
         td="target=₹1Cr, horizon=20yr, rate=12%",
         steps=(
             "Given the SIP calculator tool\n"
             "When calculate_sip(target=10000000, years=20, rate=0.12) is called\n"
             "Then monthly_sip is approximately ₹10,109"
         ),
         exp="monthly_sip ≈ 10109 (± 5% tolerance)",
         auto="Yes", cmt="FV formula: target = sip × ((1+r)^n - 1) / r"),

    dict(tc="TC_COPILOT2_003", mod="COPILOT2", req="FR-COP2-003",
         title="Multi-turn conversation maintains context",
         pri="High", sev="High",
         pre="Active LangGraph session with prior messages",
         td="2 prior messages in session",
         steps=(
             "Given a session with 2 prior exchanges\n"
             "When a follow-up question 'What about the second fund?' is sent\n"
             "Then the response references the previously discussed fund"
         ),
         exp="Response maintains conversational context; no repetition of introduction",
         auto="No", cmt="Multi-turn state management"),

    dict(tc="TC_COPILOT2_004", mod="COPILOT2", req="FR-COP2-004",
         title="Tool calling: portfolio_fetch_tool retrieves holdings",
         pri="High", sev="High",
         pre="User has holdings",
         td="td_user_ahaan",
         steps=(
             "Given the LangGraph agent with portfolio_fetch_tool available\n"
             "When a portfolio question is asked\n"
             "Then the agent calls portfolio_fetch_tool and incorporates the holdings data"
         ),
         exp="Tool call logged; response uses actual holdings data",
         auto="No", cmt=""),

    dict(tc="TC_COPILOT2_005", mod="COPILOT2", req="FR-COP2-005",
         title="Copilot V2 also blocks PAN before LLM",
         pri="Critical", sev="Critical",
         pre="PII scrubber active in V2 pipeline",
         td="td_chat_pan",
         steps=(
             "Given a V2 chat session\n"
             "When a message with PAN 'ABCDE1234F' is sent\n"
             "Then PAN is redacted before reaching OpenAI"
         ),
         exp="OpenAI request does NOT contain 'ABCDE1234F'",
         auto="Yes", cmt="PII safety must apply to V2 as well"),

    # ══════════════════════════════════════════════════════════════════════════
    # ADMIN
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_ADMIN_001", mod="ADMIN", req="FR-ADM-001",
         title="Admin can update a secret value",
         pri="Critical", sev="Critical",
         pre="Logged in as admin; secret key exists",
         td="td_user_admin; secret key='OPENAI_API_KEY'",
         steps=(
             "Given an admin user is authenticated\n"
             "When PUT /api/admin/secrets/{key} with new value is called\n"
             "Then HTTP 200 is returned and secret is updated in system_config"
         ),
         exp="HTTP 200; secret updated; masked value returned (not plaintext)",
         auto="Yes", cmt="Critical — live credential rotation"),

    dict(tc="TC_ADMIN_002", mod="ADMIN", req="FR-ADM-001",
         title="Non-admin cannot access secrets endpoint",
         pri="Critical", sev="Critical",
         pre="Logged in as role=user",
         td="td_user_ahaan",
         steps=(
             "Given a non-admin user\n"
             "When GET /api/admin/secrets is called\n"
             "Then HTTP 403 is returned"
         ),
         exp="HTTP 403",
         auto="Yes", cmt="RBAC for admin"),

    dict(tc="TC_ADMIN_003", mod="ADMIN", req="FR-ADM-002",
         title="Feature flag toggle enables/disables feature",
         pri="High", sev="High",
         pre="Feature flag 'copilot_v2' exists, currently disabled",
         td="flag=copilot_v2, enabled=true",
         steps=(
             "Given feature flag 'copilot_v2' is disabled\n"
             "When PATCH /api/admin/feature-flags/copilot_v2 with {enabled: true} is called by admin\n"
             "Then flag is enabled and Copilot V2 endpoints become accessible"
         ),
         exp="HTTP 200; flag.enabled = true; GET /api/copilot/v2/sessions returns 200 (not 404)",
         auto="Yes", cmt="Feature flag enforcement"),

    dict(tc="TC_ADMIN_004", mod="ADMIN", req="FR-ADM-003",
         title="Rules config: update action plan threshold",
         pri="High", sev="High",
         pre="Logged in as admin",
         td="param=min_quality_for_exit, value=7.0",
         steps=(
             "Given the action plan rule config\n"
             "When PATCH /api/admin/rules-config with {min_quality_for_exit: 7.0} is called\n"
             "Then subsequent plan generation uses the new threshold"
         ),
         exp="HTTP 200; config updated; next plan generation uses quality threshold 7.0",
         auto="Yes", cmt="Live rule tuning without deployment"),

    dict(tc="TC_ADMIN_005", mod="ADMIN", req="FR-ADM-004",
         title="Custom DSL rule compiles and evaluates safely",
         pri="High", sev="Critical",
         pre="Logged in as admin; safe AST evaluator running",
         td="rule: 'quality_score < 5 AND overlap_pct > 70'",
         steps=(
             "Given a valid DSL rule string\n"
             "When POST /api/admin/custom-rules with the rule is called\n"
             "Then HTTP 200 is returned with compiled rule; no arbitrary code executed"
         ),
         exp="HTTP 200; rule compiled; AST restricted to safe operators",
         auto="Yes", cmt="Safe AST — no exec/eval"),

    dict(tc="TC_ADMIN_006", mod="ADMIN", req="FR-ADM-004",
         title="Malicious DSL rule with exec() is rejected",
         pri="Critical", sev="Critical",
         pre="Logged in as admin",
         td="rule: '__import__(\"os\").system(\"rm -rf /\")'",
         steps=(
             "Given an admin attempts to inject a malicious rule with Python exec\n"
             "When POST /api/admin/custom-rules with the malicious string is called\n"
             "Then HTTP 422 is returned — unsafe AST node rejected"
         ),
         exp="HTTP 422; error = 'Unsafe expression'; no system call executed",
         auto="Yes", cmt="Security: AST sandbox must block dangerous nodes"),

    dict(tc="TC_ADMIN_007", mod="ADMIN", req="FR-ADM-005",
         title="LLM prompt sandbox executes with mock portfolio data",
         pri="Medium", sev="Medium",
         pre="Logged in as admin",
         td="Modified portfolio analysis prompt",
         steps=(
             "Given an admin edits the portfolio analysis prompt\n"
             "When POST /api/admin/prompt-sandbox/test with the prompt and mock data is called\n"
             "Then LLM response is returned without affecting production prompts"
         ),
         exp="HTTP 200; response from LLM; no production prompt modified",
         auto="No", cmt="Sandbox mode"),

    dict(tc="TC_ADMIN_008", mod="ADMIN", req="FR-ADM-006",
         title="Data pipeline monitor shows scheduled jobs",
         pri="High", sev="Medium",
         pre="APScheduler running; at least one scheduled job",
         td="Admin authenticated",
         steps=(
             "Given APScheduler is running with scheduled jobs\n"
             "When GET /api/admin/jobs is called\n"
             "Then list of jobs with last_run, next_run, status is returned"
         ),
         exp="HTTP 200; jobs array non-empty; each job has name, last_run, next_run, status",
         auto="Yes", cmt=""),

    dict(tc="TC_ADMIN_009", mod="ADMIN", req="FR-ADM-007",
         title="V3 scoring weights persist across restart",
         pri="High", sev="High",
         pre="Admin authenticated",
         td="quality_weight=0.4, health_weight=0.3",
         steps=(
             "Given an admin updates V3 scoring composite weights\n"
             "When PATCH /api/admin/v3-weights is called and service restarts\n"
             "Then POST-restart scoring uses the new weights"
         ),
         exp="Weights persisted in MongoDB; not reset on restart",
         auto="No", cmt="Persistence across restart"),

    dict(tc="TC_ADMIN_010", mod="ADMIN", req="FR-ADM-008",
         title="Admin can promote user to advisor role",
         pri="High", sev="High",
         pre="Admin authenticated; target user exists with role=user",
         td="td_user_admin; target=td_user_ahaan",
         steps=(
             "Given an admin and a target user with role=user\n"
             "When PATCH /api/admin/users/{user_id}/role with {role: 'advisor'} is called\n"
             "Then user.role = advisor"
         ),
         exp="HTTP 200; user.role = advisor",
         auto="Yes", cmt=""),

    dict(tc="TC_ADMIN_011", mod="ADMIN", req="FR-ADM-009",
         title="Whitelist management: add new email",
         pri="High", sev="High",
         pre="Admin authenticated; email not yet whitelisted",
         td="new_user@example.com",
         steps=(
             "Given an admin\n"
             "When POST /api/admin/whitelist with {email: 'new_user@example.com'} is called\n"
             "Then the email is added to whitelisted_users and the user can log in"
         ),
         exp="HTTP 201; email in whitelist; new user can authenticate",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # NIDP
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_NIDP_001", mod="NIDP", req="FR-NIDP-001",
         title="NAV ingestion job loads latest NAV from AMFI",
         pri="Critical", sev="Critical",
         pre="AMFI NAV URL reachable; job triggered manually",
         td="AMFI NAV file for current date",
         steps=(
             "Given AMFI NAV data is available\n"
             "When POST /api/admin/jobs/nav-ingest/trigger is called\n"
             "Then job runs and nav_history table is updated with today's NAVs"
         ),
         exp="job.status = completed; nav_history rows for today present",
         auto="Yes", cmt="Core data pipeline"),

    dict(tc="TC_NIDP_002", mod="NIDP", req="FR-NIDP-002",
         title="NIDP job completion emits eventType in logs",
         pri="High", sev="High",
         pre="GCP logging configured; NIDP job runs",
         td="Any NIDP job",
         steps=(
             "Given a NIDP job is triggered\n"
             "When the job completes\n"
             "Then the completion log entry contains eventType field with the job name"
         ),
         exp="GCP log entry has jsonPayload.eventType = 'nidp.{job_name}.completed'",
         auto="No", cmt="GCP metrics filter requires this field"),

    dict(tc="TC_NIDP_003", mod="NIDP", req="FR-NIDP-003",
         title="Data quality validator catches missing required fields",
         pri="High", sev="High",
         pre="DQ validation engine running",
         td="NAV record missing isin field",
         steps=(
             "Given a NAV record without the required isin field\n"
             "When the DQ validator runs\n"
             "Then a validation failure is recorded with field=isin, rule=REQUIRED"
         ),
         exp="validation_failure.field = isin; validation_failure.rule = REQUIRED",
         auto="Yes", cmt=""),

    dict(tc="TC_NIDP_004", mod="NIDP", req="FR-NIDP-004",
         title="NIDP validation_findings actual field is TEXT 'True'/'False'",
         pri="High", sev="Medium",
         pre="DQ run completed",
         td="Any completed job run",
         steps=(
             "Given a completed NIDP job with validation findings\n"
             "When querying nidp_validation_findings WHERE job_run_id = {id}\n"
             "Then finding.actual values are TEXT strings 'True' or 'False' (not booleans)"
         ),
         exp="actual IN ('True', 'False'); NOT true/false boolean",
         auto="Yes", cmt="Schema gotcha: TEXT not boolean"),

    dict(tc="TC_NIDP_005", mod="NIDP", req="FR-NIDP-005",
         title="Fund master sync enriches instrument_master from NIDP",
         pri="High", sev="High",
         pre="NIDP fund master data available",
         td="Fund master with 1000 funds",
         steps=(
             "Given NIDP fund master data is available\n"
             "When POST /api/admin/jobs/fund-master-sync/trigger is called\n"
             "Then instrument_master in PostgreSQL is updated with latest fund details"
         ),
         exp="job.status = completed; instrument_master row count increased or refreshed",
         auto="Yes", cmt=""),

    dict(tc="TC_NIDP_006", mod="NIDP", req="FR-NIDP-006",
         title="S4/S5 announcement pipeline processes NSE announcements",
         pri="High", sev="High",
         pre="S4 pipeline running; NSE announcement feed available",
         td="NSE corporate announcement",
         steps=(
             "Given the S4 pipeline is running\n"
             "When a new NSE announcement arrives\n"
             "Then it is classified, embedded (S5), and stored in documents table with pgvector"
         ),
         exp="documents table has new row; embedding non-null; classified_at present",
         auto="No", cmt="S4/S5 live as of 2026-05-07"),

    dict(tc="TC_NIDP_007", mod="NIDP", req="FR-NIDP-007",
         title="NIDP job control: all 13 jobs listable",
         pri="High", sev="Medium",
         pre="Admin authenticated",
         td="td_user_admin",
         steps=(
             "Given an admin user\n"
             "When GET /api/admin/jobs is called\n"
             "Then all 13 NIDP jobs are listed with their status"
         ),
         exp="jobs.length = 13 (or total configured); each has name, schedule, last_run",
         auto="Yes", cmt=""),

    # ══════════════════════════════════════════════════════════════════════════
    # UI
    # ══════════════════════════════════════════════════════════════════════════
    dict(tc="TC_UI_001", mod="UI", req="FR-V1F-SHELL-001",
         title="Landing page renders without error",
         pri="Critical", sev="High",
         pre="Frontend running on localhost:5173",
         td="No auth required",
         steps=(
             "Given the React application is running\n"
             "When the user navigates to /\n"
             "Then the landing page renders with Google Sign-In button visible"
         ),
         exp="HTTP 200; Google Sign-In button present; no console errors",
         auto="Yes", cmt="E2E Playwright"),

    dict(tc="TC_UI_002", mod="UI", req="FR-V1F-SHELL-001",
         title="Unauthenticated user redirected to landing from /dashboard",
         pri="Critical", sev="High",
         pre="No session cookie",
         td="No auth",
         steps=(
             "Given no active session\n"
             "When the user navigates to /dashboard\n"
             "Then they are redirected to / (landing page)"
         ),
         exp="URL = /; Google Sign-In visible",
         auto="Yes", cmt=""),

    dict(tc="TC_UI_003", mod="UI", req="FR-V1F-DASH-001",
         title="Dashboard renders portfolio summary after login",
         pri="Critical", sev="High",
         pre="User authenticated with holdings",
         td="td_user_ahaan",
         steps=(
             "Given the user is authenticated with an active portfolio\n"
             "When they navigate to /dashboard\n"
             "Then portfolio total value, today's gain/loss, and fund count are displayed"
         ),
         exp="Total value visible; gain/loss with colour coding; at least 1 fund card",
         auto="Yes", cmt="Playwright E2E"),

    dict(tc="TC_UI_004", mod="UI", req="FR-V1F-PORT-001",
         title="CAS upload UI shows progress and success toast",
         pri="Critical", sev="High",
         pre="User on portfolio page",
         td="td_cas_nsdl",
         steps=(
             "Given the user is on the Portfolio page\n"
             "When they upload a valid CAS PDF\n"
             "Then a progress indicator appears and then a success toast with holdings count"
         ),
         exp="Progress spinner visible; success toast says 'X holdings imported'",
         auto="Yes", cmt="Playwright E2E upload flow"),

    dict(tc="TC_UI_005", mod="UI", req="FR-V1F-PORT-002",
         title="Holdings table displays all uploaded holdings",
         pri="Critical", sev="High",
         pre="CAS uploaded with 15 holdings",
         td="td_cas_nsdl",
         steps=(
             "Given a portfolio with 15 holdings imported from CAS\n"
             "When the Portfolio page is viewed\n"
             "Then the holdings table shows 15 rows with fund name, value, gain/loss"
         ),
         exp="Table has 15 data rows; fund names match CAS; values non-zero",
         auto="Yes", cmt=""),

    dict(tc="TC_UI_006", mod="UI", req="FR-V2F-PLAN-001",
         title="Plan board renders action cards in priority order",
         pri="Critical", sev="High",
         pre="Active plan with ≥3 actions",
         td="td_user_ahaan with plan",
         steps=(
             "Given an active action plan with 3+ actions\n"
             "When the Plan Board page is loaded\n"
             "Then action cards appear in priority order (Critical first)"
         ),
         exp="First card has priority=Critical; cards sorted by priority desc",
         auto="Yes", cmt=""),

    dict(tc="TC_UI_007", mod="UI", req="FR-V2F-PLAN-002",
         title="Mark action as done updates plan progress",
         pri="High", sev="High",
         pre="Active plan with pending actions",
         td="td_user_ahaan",
         steps=(
             "Given an action card in pending state\n"
             "When the user clicks 'Mark Done'\n"
             "Then the card moves to done state and progress % increases"
         ),
         exp="action.status = done; progress_pct > previous value",
         auto="Yes", cmt=""),

    dict(tc="TC_UI_008", mod="UI", req="FR-V2F-INS-001",
         title="V2 insights layer shows overlap heatmap",
         pri="High", sev="Medium",
         pre="Portfolio with overlapping funds",
         td="td_portfolio_overlap",
         steps=(
             "Given a portfolio with overlapping funds\n"
             "When the Insights page is viewed\n"
             "Then an overlap heatmap is rendered showing fund pair overlap percentages"
         ),
         exp="Heatmap visible; overlapping cells highlighted; values match API response",
         auto="Yes", cmt=""),

    dict(tc="TC_UI_009", mod="UI", req="FR-V1F-MKT-001",
         title="Market dashboard shows indices with cache refresh",
         pri="Medium", sev="Low",
         pre="Market data available",
         td="None",
         steps=(
             "Given the Market Dashboard page\n"
             "When the page loads\n"
             "Then NIFTY 50, SENSEX, and top sector indices are displayed with % change"
         ),
         exp="Indices visible; % change colour coded green/red; refresh indicator present",
         auto="Yes", cmt="Cached data — not real-time"),

    dict(tc="TC_UI_010", mod="UI", req="FR-V1F-SHELL-002",
         title="INR currency formatter renders ₹ sign correctly",
         pri="Low", sev="Low",
         pre="Any page with currency values",
         td="Amount: 1500000",
         steps=(
             "Given any page displaying a currency amount of ₹15,00,000\n"
             "When the page renders\n"
             "Then the value is displayed as '₹15,00,000' in Indian number format"
         ),
         exp="Display = '₹15,00,000' (Indian lakh-crore format, not Western)",
         auto="Yes", cmt="INR formatting"),

]

# ── Requirement → TC mapping (for RTM) ───────────────────────────────────────
def build_rtm(test_cases):
    rtm = {}
    for tc in test_cases:
        req = tc["req_id"]
        if req not in rtm:
            rtm[req] = []
        rtm[req].append(tc["tc_id"])
    return rtm


# ── openpyxl helpers ──────────────────────────────────────────────────────────
def header_fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header_row(ws, headers, col_widths, bg="1A237E", fg="FFFFFF"):
    hdr_font  = Font(bold=True, color=fg, size=10)
    hdr_fill  = header_fill(bg)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def priority_colour(priority):
    p = priority.lower()
    if p == "critical": return COLOUR["critical"]
    if p == "high":     return COLOUR["high"]
    if p == "medium":   return COLOUR["medium"]
    return COLOUR["low"]

def write_tc_row(ws, row_num, tc):
    is_alt = row_num % 2 == 0
    alt_fill = header_fill(COLOUR["alt_row"]) if is_alt else None
    pri_fill  = header_fill(priority_colour(tc["priority"]))

    values = [
        tc["tc_id"], tc["module"], tc["req_id"], tc["title"],
        tc["priority"], tc["severity"],
        tc["preconditions"], tc["test_data"], tc["steps"],
        tc["expected"], "", "Not Run",
        tc["automation"], tc["comments"],
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border    = thin_border()
        cell.font      = Font(size=9)
        if alt_fill:
            cell.fill = alt_fill
        if col_idx == 5:   # Priority column — colour by level
            cell.fill = pri_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
    ws.row_dimensions[row_num].height = 72


# ── Main builder ──────────────────────────────────────────────────────────────
def normalise(raw_list):
    """Convert list-of-dicts shorthand to full field names."""
    out = []
    for r in raw_list:
        out.append({
            "tc_id":        r["tc"],
            "module":       r["mod"],
            "req_id":       r["req"],
            "title":        r["title"],
            "priority":     r["pri"],
            "severity":     r["sev"],
            "preconditions": r["pre"],
            "test_data":    r["td"],
            "steps":        r["steps"],
            "expected":     r["exp"],
            "automation":   r["auto"],
            "comments":     r["cmt"],
        })
    return out


def build_workbook(test_cases):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    tcs = normalise(test_cases)

    # ── Sheet 1: Functional Test Cases ───────────────────────────────────────
    ws_tc = wb.create_sheet("Functional Test Cases")
    apply_header_row(ws_tc, HEADERS, COL_WIDTHS)
    for i, tc in enumerate(tcs, start=2):
        write_tc_row(ws_tc, i, tc)

    # ── Sheet 2: RTM ─────────────────────────────────────────────────────────
    ws_rtm = wb.create_sheet("RTM")
    rtm_headers = ["Req ID", "Module", "Test Cases", "Coverage Count", "Status"]
    rtm_widths   = [20, 14, 50, 16, 14]
    apply_header_row(ws_rtm, rtm_headers, rtm_widths, bg="004D40")

    rtm = build_rtm(tcs)
    # Gather module from test cases
    req_module = {tc["req_id"]: tc["module"] for tc in tcs}
    for row_num, (req_id, tc_ids) in enumerate(sorted(rtm.items()), start=2):
        ws_rtm.cell(row=row_num, column=1, value=req_id).border = thin_border()
        ws_rtm.cell(row=row_num, column=2, value=req_module.get(req_id, "")).border = thin_border()
        ws_rtm.cell(row=row_num, column=3, value=", ".join(tc_ids)).border = thin_border()
        ws_rtm.cell(row=row_num, column=4, value=len(tc_ids)).border = thin_border()
        ws_rtm.cell(row=row_num, column=5, value="Covered").border = thin_border()
        ws_rtm.row_dimensions[row_num].height = 20

    # ── Sheet 3: Automation Candidates ───────────────────────────────────────
    ws_auto = wb.create_sheet("Automation Candidates")
    auto_headers = ["TC ID", "Module", "Title", "Priority", "Why Automate", "Playwright Pattern"]
    auto_widths   = [14, 12, 40, 10, 35, 45]
    apply_header_row(ws_auto, auto_headers, auto_widths, bg="4A148C")

    playwright_patterns = {
        "AUTH":       "test('TC — {title}', async ({ request }) => { const r = await request.post('/api/auth/google', ...) })",
        "PORTFOLIO":  "test('TC — {title}', async ({ request }) => { await request.post('/api/portfolios/...', ...) })",
        "TAX":        "test('TC — {title}', async ({ request }) => { const r = await request.get('/api/portfolios/{id}/tax-summary') })",
        "PLAN":       "test('TC — {title}', async ({ request }) => { await request.post('/api/plans/generate', ...) })",
        "V3SCORE":    "test('TC — {title}', async ({ request }) => { const r = await request.get('/api/v3/score/{isin}') })",
        "COPILOT":    "test('TC — {title}', async ({ request }) => { await request.post('/api/copilot/sessions/{id}/chat', ...) })",
        "ADMIN":      "test('TC — {title}', async ({ request }) => { await request.put('/api/admin/...', ...) })",
        "UI":         "test('TC — {title}', async ({ page }) => { await page.goto('/dashboard'); await expect(page.locator(...)).toBeVisible() })",
    }

    auto_row = 2
    for tc in tcs:
        if tc["automation"] == "Yes":
            mod = tc["module"].split("2")[0]  # COPILOT2 → COPILOT
            pattern = playwright_patterns.get(mod, "test('{tc_id}', async ({ request }) => { ... })")
            why = f"Stable API contract; deterministic output" if tc["priority"] == "Critical" else "Repeatable, no external dependency"
            ws_auto.cell(row=auto_row, column=1, value=tc["tc_id"]).border = thin_border()
            ws_auto.cell(row=auto_row, column=2, value=tc["module"]).border = thin_border()
            ws_auto.cell(row=auto_row, column=3, value=tc["title"]).border = thin_border()
            ws_auto.cell(row=auto_row, column=4, value=tc["priority"]).border = thin_border()
            ws_auto.cell(row=auto_row, column=5, value=why).border = thin_border()
            ws_auto.cell(row=auto_row, column=6, value=pattern.format(title=tc["title"], tc_id=tc["tc_id"])).border = thin_border()
            ws_auto.row_dimensions[auto_row].height = 32
            auto_row += 1

    # ── Sheet 4: Module Summary ───────────────────────────────────────────────
    ws_sum = wb.create_sheet("Module Summary")
    sum_headers = ["Module", "Total TCs", "Critical", "High", "Medium", "Low",
                   "Automation Candidates", "Coverage %"]
    sum_widths   = [14, 10, 10, 10, 10, 10, 20, 12]
    apply_header_row(ws_sum, sum_headers, sum_widths, bg="B71C1C")

    from collections import defaultdict, Counter
    module_tcs = defaultdict(list)
    for tc in tcs:
        module_tcs[tc["module"]].append(tc)

    for row_num, (mod, mod_tcs) in enumerate(sorted(module_tcs.items()), start=2):
        prio_counts = Counter(tc["priority"] for tc in mod_tcs)
        auto_count  = sum(1 for tc in mod_tcs if tc["automation"] == "Yes")
        total       = len(mod_tcs)
        cov_pct     = round((auto_count / total) * 100) if total else 0
        ws_sum.cell(row=row_num, column=1, value=mod)
        ws_sum.cell(row=row_num, column=2, value=total)
        ws_sum.cell(row=row_num, column=3, value=prio_counts.get("Critical", 0))
        ws_sum.cell(row=row_num, column=4, value=prio_counts.get("High", 0))
        ws_sum.cell(row=row_num, column=5, value=prio_counts.get("Medium", 0))
        ws_sum.cell(row=row_num, column=6, value=prio_counts.get("Low", 0))
        ws_sum.cell(row=row_num, column=7, value=auto_count)
        ws_sum.cell(row=row_num, column=8, value=f"{cov_pct}%")
        for col in range(1, 9):
            ws_sum.cell(row=row_num, column=col).border = thin_border()
        ws_sum.row_dimensions[row_num].height = 20

    # Totals row
    total_row = len(module_tcs) + 2
    ws_sum.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=total_row, column=2, value=len(tcs)).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=3, value=sum(1 for tc in tcs if tc["priority"] == "Critical")).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=4, value=sum(1 for tc in tcs if tc["priority"] == "High")).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=5, value=sum(1 for tc in tcs if tc["priority"] == "Medium")).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=6, value=sum(1 for tc in tcs if tc["priority"] == "Low")).font = Font(bold=True)
    ws_sum.cell(row=total_row, column=7, value=sum(1 for tc in tcs if tc["automation"] == "Yes")).font = Font(bold=True)

    return wb, tcs


def write_csv(tcs, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "tc_id", "module", "req_id", "title", "priority", "severity",
            "preconditions", "test_data", "steps", "expected",
            "actual_result", "status", "automation", "comments",
        ])
        writer.writeheader()
        for tc in tcs:
            writer.writerow({**tc, "actual_result": "", "status": "Not Run"})


def write_rtm_csv(tcs, path):
    rtm = build_rtm(tcs)
    req_module = {tc["req_id"]: tc["module"] for tc in tcs}
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Req ID", "Module", "Test Cases", "Coverage Count"])
        for req_id, tc_ids in sorted(rtm.items()):
            writer.writerow([req_id, req_module.get(req_id, ""), ", ".join(tc_ids), len(tc_ids)])


if __name__ == "__main__":
    wb, tcs = build_workbook(TEST_CASES)

    xlsx_path = os.path.join(OUT_DIR, "functional_test_cases.xlsx")
    csv_path  = os.path.join(OUT_DIR, "functional_test_cases.csv")
    rtm_path  = os.path.join(OUT_DIR, "requirement_traceability_matrix.csv")

    wb.save(xlsx_path)
    write_csv(tcs, csv_path)
    write_rtm_csv(tcs, rtm_path)

    from collections import Counter
    prio_counts = Counter(tc["priority"] for tc in tcs)
    auto_count  = sum(1 for tc in tcs if tc["automation"] == "Yes")

    print(f"Generated {len(tcs)} test cases")
    print(f"  Critical: {prio_counts['Critical']}  High: {prio_counts['High']}  "
          f"Medium: {prio_counts['Medium']}  Low: {prio_counts['Low']}")
    print(f"  Automation candidates: {auto_count}/{len(tcs)}")
    print(f"  {xlsx_path}")
    print(f"  {csv_path}")
    print(f"  {rtm_path}")
