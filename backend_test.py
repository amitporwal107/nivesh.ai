import requests
import sys
import json
import io
from datetime import datetime

class WealthSystemAPITester:
    def __init__(self, base_url="https://ai-advisor-30.preview.emergentagent.com"):
        self.base_url = base_url
        self.session_token = "test_session_wealth001"  # Pre-created test session
        self.tests_run = 0
        self.tests_passed = 0
        self.failed_tests = []

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None):
        """Run a single API test"""
        url = f"{self.base_url}/api/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        
        # Add session token via Authorization header
        if self.session_token:
            headers['Authorization'] = f'Bearer {self.session_token}'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        print(f"   URL: {url}")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    # Remove Content-Type for file uploads
                    headers.pop('Content-Type', None)
                    response = requests.post(url, files=files, headers=headers)
                else:
                    response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                try:
                    response_data = response.json()
                    print(f"   Response: {json.dumps(response_data, indent=2)[:200]}...")
                except:
                    print(f"   Response: {response.text[:100]}...")
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                print(f"   Response: {response.text[:200]}...")
                self.failed_tests.append({
                    "test": name,
                    "expected": expected_status,
                    "actual": response.status_code,
                    "response": response.text[:200]
                })

            return success, response.json() if success and response.text else {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            self.failed_tests.append({
                "test": name,
                "error": str(e)
            })
            return False, {}

    def test_auth_me(self):
        """Test /auth/me endpoint"""
        success, response = self.run_test(
            "Get Current User",
            "GET",
            "auth/me",
            200
        )
        if success and 'user_id' in response:
            print(f"   User: {response.get('name')} ({response.get('email')})")
            return response
        return None

    def test_auth_logout(self):
        """Test logout endpoint"""
        success, _ = self.run_test(
            "Logout",
            "POST",
            "auth/logout",
            200
        )
        return success

    def test_get_holdings(self):
        """Test get holdings"""
        success, response = self.run_test(
            "Get Holdings",
            "GET",
            "portfolio/holdings",
            200
        )
        if success:
            print(f"   Found {len(response)} holdings")
            return response
        return []

    def test_add_holding(self):
        """Test add holding"""
        test_holding = {
            "name": "Test Stock SDET",
            "ticker": "TESTSDET",
            "asset_type": "equity",
            "quantity": 10,
            "buy_price": 1000,
            "current_price": 1200,
            "sector": "IT",
            "buy_date": "2024-01-01"
        }
        
        success, response = self.run_test(
            "Add Holding",
            "POST",
            "portfolio/holdings",
            201,
            data=test_holding
        )
        return response.get('holding_id') if success else None

    def test_update_holding(self, holding_id):
        """Test update holding"""
        update_data = {
            "current_price": 1300,
            "quantity": 15
        }
        
        success, response = self.run_test(
            "Update Holding",
            "PUT",
            f"portfolio/holdings/{holding_id}",
            200,
            data=update_data
        )
        return success

    def test_delete_holding(self, holding_id):
        """Test delete holding"""
        success, _ = self.run_test(
            "Delete Holding",
            "DELETE",
            f"portfolio/holdings/{holding_id}",
            200
        )
        return success

    def test_csv_upload(self):
        """Test CSV upload via new unified endpoint"""
        csv_content = """Name,Ticker,Type,Quantity,Buy Price,Current Price,Sector
Test CSV Stock,TESTCSV,equity,5,800,900,Banking
Test MF,TESTMF,mutual_fund,100,50,55,Financial Services"""
        
        csv_file = io.StringIO(csv_content)
        files = {'file': ('test_holdings.csv', csv_file.getvalue(), 'text/csv')}
        
        success, response = self.run_test(
            "CSV Upload (New Endpoint)",
            "POST",
            "portfolio/upload",
            200,
            files=files
        )
        if success:
            print(f"   Imported {response.get('count', 0)} holdings")
        return success

    def test_excel_upload(self):
        """Test Excel upload via unified endpoint"""
        try:
            import openpyxl
            from io import BytesIO
            
            # Create Excel file in memory
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Headers
            ws['A1'] = 'Name'
            ws['B1'] = 'Ticker'
            ws['C1'] = 'Type'
            ws['D1'] = 'Quantity'
            ws['E1'] = 'Buy Price'
            ws['F1'] = 'Current Price'
            ws['G1'] = 'Sector'
            
            # Data rows
            ws['A2'] = 'Test Excel Stock'
            ws['B2'] = 'TESTXL'
            ws['C2'] = 'equity'
            ws['D2'] = 8
            ws['E2'] = 1200
            ws['F2'] = 1350
            ws['G2'] = 'IT'
            
            ws['A3'] = 'Test Excel MF'
            ws['B3'] = 'TESTXLMF'
            ws['C3'] = 'mutual_fund'
            ws['D3'] = 200
            ws['E3'] = 25
            ws['F3'] = 28
            ws['G3'] = 'Financial Services'
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            files = {'file': ('test_holdings.xlsx', excel_buffer.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            
            success, response = self.run_test(
                "Excel Upload",
                "POST",
                "portfolio/upload",
                200,
                files=files
            )
            if success:
                print(f"   Imported {response.get('count', 0)} holdings")
            return success
            
        except ImportError:
            print("❌ openpyxl not available - skipping Excel test")
            return False
        except Exception as e:
            print(f"❌ Excel test failed: {e}")
            return False

    def test_cas_pdf_upload(self):
        """Test CAS PDF async upload (mock PDF)"""
        # Create a simple mock PDF content
        mock_pdf_content = b"%PDF-1.4\n1 0 obj\n<<\n/Type /Catalog\n/Pages 2 0 R\n>>\nendobj\n2 0 obj\n<<\n/Type /Pages\n/Kids [3 0 R]\n/Count 1\n>>\nendobj\n3 0 obj\n<<\n/Type /Page\n/Parent 2 0 R\n/MediaBox [0 0 612 792]\n>>\nendobj\nxref\n0 4\n0000000000 65535 f \n0000000009 00000 n \n0000000074 00000 n \n0000000120 00000 n \ntrailer\n<<\n/Size 4\n/Root 1 0 R\n>>\nstartxref\n179\n%%EOF"
        
        files = {'file': ('test_cas.pdf', mock_pdf_content, 'application/pdf')}
        
        success, response = self.run_test(
            "CAS PDF Upload (Async)",
            "POST",
            "portfolio/upload",
            200,
            files=files
        )
        
        if success:
            task_id = response.get('task_id')
            status = response.get('status')
            print(f"   Task ID: {task_id}")
            print(f"   Status: {status}")
            
            if task_id and status == 'processing':
                return task_id
        return None

    def test_upload_status_polling(self, task_id):
        """Test upload status polling"""
        if not task_id:
            print("❌ No task_id provided for status polling")
            return False
            
        success, response = self.run_test(
            "Upload Status Polling",
            "GET",
            f"portfolio/upload-status/{task_id}",
            200
        )
        
        if success:
            status = response.get('status', 'unknown')
            message = response.get('message', '')
            count = response.get('count', 0)
            print(f"   Status: {status}")
            print(f"   Message: {message}")
            print(f"   Count: {count}")
            return True
        return False

    def test_portfolio_analytics(self):
        """Test portfolio analytics with new dashboard fields"""
        success, response = self.run_test(
            "Portfolio Analytics",
            "GET",
            "portfolio/analytics",
            200
        )
        if success:
            print(f"   Total Value: ₹{response.get('current_value', 0)}")
            print(f"   Returns: ₹{response.get('total_returns', 0)} ({response.get('returns_pct', 0):.1f}%)")
            print(f"   Risk Score: {response.get('risk_score', 0)} ({response.get('risk_label', 'N/A')})")
            print(f"   Holdings Count: {response.get('holdings_count', 0)}")
            
            # Test new dashboard fields
            required_fields = ['performance_trend', 'heatmap_data', 'day_change', 'day_change_pct']
            missing_fields = []
            
            for field in required_fields:
                if field not in response:
                    missing_fields.append(field)
                else:
                    print(f"   ✅ {field}: {type(response[field])} with {len(response[field]) if isinstance(response[field], list) else 'N/A'} items")
            
            if missing_fields:
                print(f"   ❌ Missing required fields: {missing_fields}")
                return False
                
            # Validate performance_trend structure
            if response.get('performance_trend'):
                trend_sample = response['performance_trend'][0] if response['performance_trend'] else {}
                if 'date' in trend_sample and 'value' in trend_sample:
                    print(f"   ✅ Performance trend has correct structure")
                else:
                    print(f"   ❌ Performance trend missing date/value fields")
                    return False
            
            # Validate heatmap_data structure
            if response.get('heatmap_data'):
                heatmap_sample = response['heatmap_data'][0] if response['heatmap_data'] else {}
                required_heatmap_fields = ['name', 'value', 'return_pct']
                for field in required_heatmap_fields:
                    if field not in heatmap_sample:
                        print(f"   ❌ Heatmap data missing {field} field")
                        return False
                print(f"   ✅ Heatmap data has correct structure")
            
            # Validate day_change fields
            if 'day_change' in response and 'day_change_pct' in response:
                print(f"   ✅ Day change: {response['day_change']} ({response['day_change_pct']}%)")
            else:
                print(f"   ❌ Day change fields missing")
                return False
                
        return success

    def test_chat_messages(self):
        """Test get chat messages"""
        success, response = self.run_test(
            "Get Chat Messages",
            "GET",
            "chat/messages",
            200
        )
        if success:
            print(f"   Found {len(response)} messages")
        return success

    def test_send_chat(self):
        """Test send chat message"""
        test_message = {
            "message": "What is my portfolio's risk level?"
        }
        
        success, response = self.run_test(
            "Send Chat Message",
            "POST",
            "chat/send",
            200,
            data=test_message
        )
        if success:
            print(f"   AI Response: {response.get('ai_message', {}).get('content', '')[:100]}...")
        return success

    def test_clear_chat(self):
        """Test clear chat"""
        success, _ = self.run_test(
            "Clear Chat",
            "DELETE",
            "chat/clear",
            200
        )
        return success

    def test_get_insights(self):
        """Test get insights"""
        success, response = self.run_test(
            "Get Insights",
            "GET",
            "insights",
            200
        )
        if success:
            print(f"   Found {len(response)} insights")
        return success

    def test_generate_insights(self):
        """Test generate insights"""
        success, response = self.run_test(
            "Generate Insights",
            "POST",
            "insights/generate",
            200
        )
        if success:
            insights = response.get('insights', [])
            print(f"   Generated {len(insights)} insights")
            for insight in insights[:2]:  # Show first 2
                print(f"   - {insight.get('title', '')}: {insight.get('type', '')} ({insight.get('priority', '')})")
        return success

def main():
    print("🚀 Starting Agentic Wealth System API Tests")
    print("=" * 60)
    
    tester = WealthSystemAPITester()
    
    # Test authentication
    print("\n📋 AUTHENTICATION TESTS")
    user = tester.test_auth_me()
    if not user:
        print("❌ Authentication failed - cannot proceed with other tests")
        return 1
    
    # Test portfolio endpoints
    print("\n📊 PORTFOLIO TESTS")
    holdings = tester.test_get_holdings()
    
    # Test CRUD operations
    holding_id = tester.test_add_holding()
    if holding_id:
        tester.test_update_holding(holding_id)
        # Don't delete immediately, keep for other tests
    
    # Test new upload functionality
    print("\n📤 UPLOAD TESTS")
    tester.test_csv_upload()
    tester.test_excel_upload()
    
    # Test CAS PDF async upload
    task_id = tester.test_cas_pdf_upload()
    if task_id:
        tester.test_upload_status_polling(task_id)
    
    tester.test_portfolio_analytics()
    
    # Test AI features
    print("\n🤖 AI FEATURES TESTS")
    tester.test_chat_messages()
    tester.test_send_chat()
    tester.test_get_insights()
    tester.test_generate_insights()
    
    # Test cleanup
    print("\n🧹 CLEANUP TESTS")
    tester.test_clear_chat()
    if holding_id:
        tester.test_delete_holding(holding_id)
    
    tester.test_auth_logout()
    
    # Print results
    print("\n" + "=" * 60)
    print(f"📊 FINAL RESULTS")
    print(f"Tests passed: {tester.tests_passed}/{tester.tests_run}")
    print(f"Success rate: {(tester.tests_passed/tester.tests_run*100):.1f}%")
    
    if tester.failed_tests:
        print(f"\n❌ FAILED TESTS:")
        for failure in tester.failed_tests:
            print(f"  - {failure.get('test', 'Unknown')}: {failure}")
    
    return 0 if tester.tests_passed == tester.tests_run else 1

if __name__ == "__main__":
    sys.exit(main())